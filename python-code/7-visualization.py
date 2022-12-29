#!/usr/bin/env python
# coding: utf-8

# In[20]:


import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from scipy import stats
import os
import openpyxl
import graphviz  
from sklearn import tree
from sklearn.tree import DecisionTreeRegressor
from sklearn.tree import plot_tree
import matplotlib.pyplot as plt
#from IPython.display import Image  
from sklearn.tree import export_graphviz
import warnings
warnings.filterwarnings('ignore')


# In[21]:

path_figure = '/Figure/figurefile/'
data_path2 = '/response/response_original'
data_path ='/response/response_netcomi'
###################################################################################
path_original ='/Figure/RF/response_original/'
path_aug ='/response/response_aug/'
path_Disease='/Figure/RF/response_Disease/'
####################################################################################
path_BNN = '/Figure/BNN/ALL_response/'
path_aug_BNN ='/Figure/BNN/response_aug/'
path_Disease_BNN='/Figure/BNN/Disease_response/'

hue_order = ['Phylum','Class','Order','Family','Genus']
list_level = ['Class', 'Family', 'Genus', 'Order', 'Phylum']
order= ['Yield_Plant','Yield_Meter', 'Black_Scurf','Scab','Scabpit','Scabsuper']


# # 2-ML&NetComi

# In[22]:


path_list2 = []
#reading files in folder response
for root, dirs, files in os.walk(data_path2, topdown=False):
    for path in dirs:
        path_list2.append(path)
#reading sheet name       
wb = openpyxl.load_workbook(data_path2+'/'+path+'/feature_selection.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list2, index=range(0,700))
    for folder in path_list2:
        data_temp = pd.read_excel(data_path2+'/'+folder+'/feature_selection.xlsx', sheet_name=sheet_name)
        temp_df[folder].iloc[
            range(0, len(data_temp['Unnamed: 0'].values))] = data_temp['Unnamed: 0'].values

    results_dic[sheet_name] = temp_df
    
for key in results_dic.keys():
    results_dic[key] = results_dic[key].iloc[range(0,max(30, round(len(data_temp)/3)))]


# In[23]:


path_list = []

for root, dirs, files in os.walk(data_path, topdown=False):
    for path in dirs:
        path_list.append(path)


# In[24]:


all_data = dict.fromkeys(dirs)

for folder in path_list:
    
    all_data[folder] = dict.fromkeys(list_level)
    temp_path = data_path+'/'+folder
    temp_file_list = os.listdir(temp_path)
    for file in temp_file_list:
        if 'Count' in file:
            level = file.split('_')[-1][:-4]
            data_temp = pd.read_csv(data_path+'/'+folder+'/'+file,index_col=0)
            data_temp.sort_values(by='abs.diff..x', ascending=False, inplace=True)
            all_data[folder][level] = data_temp.iloc[range(0,max(30, round(len(data_temp)/3)))]

print('Done.')


# In[25]:


for col in range(0,len(list_level)):
    level = list_level[col]
    print(level)
    response_list = path_list
    feature_list = []
    for response in response_list:
        feature_list.append(all_data[response][level].index.union(results_dic[level][response][results_dic[level][response].notnull()].values))
    feature_list = [item for subitem in feature_list for item in subitem]
    feature_list = np.unique(feature_list)
    print(len(feature_list))
    matrix_df = pd.DataFrame(columns = response_list, index = feature_list)
    for response in response_list:
        for feature in feature_list:
                if (feature in all_data[response][level].index) & (feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 3
                elif(feature in all_data[response][level].index) &(feature not in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 2##NetComi
                elif (feature not in all_data[response][level].index) &(feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 1##ML
                else:
                    matrix_df[response].loc[feature] = 0##NotMLnotNetcomi
    matrix_df['Sum'] = 0
    for index in matrix_df.index:
        matrix_df['Sum'].loc[index] = sum(matrix_df.loc[index].values)
    
    matrix_df.sort_values(by='Sum', ascending=False, inplace=True)
    ################################
    matrix_df = matrix_df[['Yield_Plant','Yield_Meter', 'Black_Scurf','Scab','Scabpit','Scabsuper','Sum']]
    
    matrix_df=matrix_df.iloc[0:30,:]
    plt.figure(figsize=(9,13), dpi=600)
    sns.heatmap(matrix_df.drop(columns='Sum').astype('float'),
           linewidths=.1, vmin=0, vmax=3,annot=True, cbar=False,cmap="Blues")
    
    plt.title(level)
    plt.savefig(path_figure+level+'Feautureselection.png',bbox_inches = 'tight',dpi=600,pad_inches=0,facecolor='white')


# # 4-ALL-RF

# In[26]:


file_list_response=[]
for file_response in os.listdir(path_original):
    if (file_response != '.DS_Store') & (file_response != 'Icon'): 
        print(file_response)
        file_list_response.append(file_response)  
        path_r= path_original+'/'+file_response
        file_list=[]
        for file in os.listdir(path_r):  
            if (file != '.DS_Store')& (file != 'Icon\r'): 
                if file=='classification_RF.xlsx': 
                    print(file)
                    os.rename(path_r+'/'+'classification_RF.xlsx', path_r+'/'+'1.ALL-OTU.xlsx')
                elif file=='classification_RF_FS_notImportant.xlsx': 
                    print(file)
                    os.rename(path_r+'/'+'classification_RF_FS_notImportant.xlsx',path_r+'/'+'2.OTU-Score0.xlsx')
                elif file=='classification_RF_FS_ML.xlsx': 
                    print(file)
                    os.rename(path_r+'/'+'classification_RF_FS_ML.xlsx',path_r+'/'+'3.OTU-Score1.xlsx')

                elif file=='classification_RF_FS_Netcomi.xlsx': 
                    print(file)
                    os.rename(path_r+'/'+'classification_RF_FS_Netcomi.xlsx', path_r+'/'+'4.OTU-Score2.xlsx')
                elif file=='classification_RF_FS_score3.xlsx': 
                    print(file)
                    os.rename(path_r+'/'+'classification_RF_FS_score3.xlsx', path_r+'/'+'5.OTU-Score3.xlsx')

                    
                ###############################################    
                elif file=='classification_RF_alpha_diversity.xlsx': 
                    print(file)
                    os.rename(path_r+'/'+'classification_RF_alpha_diversity.xlsx', path_r+'/'+'6.Alpha.xlsx')
                elif file=='classification_RF_soil_chemistry_2.xlsx': 
                    print(file)
                    os.rename(path_r+'/'+'classification_RF_soil_chemistry_2.xlsx', path_r+'/'+'7.Soil.xlsx')
                elif file=='classification_RF_disease_suppression_2.xlsx': 
                    print(file)
                    os.rename(path_r+'/'+'classification_RF_disease_suppression_2.xlsx', path_r+'/'+'8.Disease.xlsx')
                elif file=='classification_alpha+soil.xlsx': 
                    print(file)
                    os.rename(path_r+'/'+'classification_alpha+soil.xlsx', path_r+'/'+'9.Alpha+Soil.xlsx')
                elif file=='classification_RF+soil+disease.xlsx': 
                    print(file)
                    os.rename(path_r+'/'+'classification_RF+soil+disease.xlsx', path_r+'/'+'10.Soil+Disease.xlsx')
                elif file=='classification_alpha+soil+disease.xlsx': 
                    print(file)
                    os.rename(path_r+'/'+'classification_alpha+soil+disease.xlsx', path_r+'/'+'11.Alpha+Soil+Disease.xlsx')
                    
                 ############################################### 
                 
                elif file=='classification_RF_FS_soil_chemistry.xlsx': 
                    print(file)
                    os.rename(path_r+'/'+'classification_RF_FS_soil_chemistry.xlsx', path_r+'/'+'12.OTU-Score3+Soil.xlsx')
                elif file=='classification_RF_FS_disease_suppression.xlsx': 
                    print(file)
                    os.rename(path_r+'/'+'classification_RF_FS_disease_suppression.xlsx',path_r+'/'+'13.OTU-Score3+Disease.xlsx')
                elif file=='classification_important_otu+soil+disease.xlsx': 
                    print(file)
                    os.rename(path_r+'/'+'classification_important_otu+soil+disease.xlsx', path_r+'/'+'14.OTU-Score3+Soil+Disease.xlsx')                   
                    


# In[27]:


for file_response in os.listdir(path_original):
    if (file_response != '.DS_Store') & (file_response != 'Icon'): 
        print(file_response)
        file_list_response.append(file_response)  
        path_r= path_original+file_response
        file_list=[]
        os.chdir(path_r)
        for file in os.listdir(path_r):  
            if (file != '.DS_Store')& (file != 'Icon\r'): 
                if file=='10.Soil+Disease.xlsx': 
                    print(file)
                    os.rename('10.Soil+Disease.xlsx', '91.Soil+Disease.xlsx')
                elif file=='11.Alpha+Soil+Disease.xlsx': 
                    print(file)
                    os.rename('11.Alpha+Soil+Disease.xlsx', '92.Alpha+Soil+Disease.xlsx')
                elif file=='12.OTU-Score3+Soil.xlsx': 
                    print(file)
                    os.rename('12.OTU-Score3+Soil.xlsx', '93.OTU-Score3+Soil.xlsx')
                elif file=='13.OTU-Score3+Disease.xlsx': 
                    print(file)
                    os.rename('13.OTU-Score3+Disease.xlsx', '94.OTU-Score3+Disease.xlsx')
                elif file=='14.OTU-Score3+Soil+Disease.xlsx': 
                    print(file)
                    os.rename('14.OTU-Score3+Soil+Disease.xlsx', '95.OTU-Score3+Soil+Disease.xlsx')


# In[28]:


ResultFinal=pd.DataFrame()
file_list_response=[]
for file_response in sorted(os.listdir(path_original)):
    if (file_response != '.DS_Store') & (file_response != 'Icon'): 
        print(file_response)
        file_list_response.append(file_response)  
        path_r= path_original+file_response
        Result=pd.DataFrame()
        file_list=[]
        for file in sorted(os.listdir(path_r)):  
            if (file != '.DS_Store')& (file != 'Icon\r'):   
                file_list.append(file) 
                wb = openpyxl.load_workbook(path_original+file_response+'/'+file)
                sheet_list = wb.sheetnames
                
                #print(sheet_list)
                ALL_value = []
                for sheet in range(1,len(sheet_list)+1):
                    #print(sheet-1)
                    response = pd.read_excel(path_original+file_response+'/'+file, sheet_name=sheet_list[sheet-1])
                    ALL_value.append(response.iloc[4,1:].values)
                CC= [file ,ALL_value]
        
                Result = pd.concat( [Result, pd.DataFrame(CC).transpose()],axis=0) 
        Result.columns=[file, file_response]
        print(Result.shape)     
        ResultFinal=  pd.concat([ResultFinal, pd.DataFrame(Result)],axis=1)


# In[29]:


PROPS = {
'boxprops':{'facecolor':'none', 'edgecolor':'blue'},
'medianprops':{'color':'black'},
'whiskerprops':{'color':'blue'}}


# In[30]:


n_model=ResultFinal.shape[0]
file_list_response= ['Yield_Plant','Yield_Meter', 'Black_Scurf','Scab','Scabpit','Scabsuper']
EE = ResultFinal[file_list_response]
file_list = [s.replace('.xlsx', '') for s in file_list]
file_list = [s.replace('91.', '') for s in file_list]
file_list = [s.replace('92.', '') for s in file_list]
file_list = [s.replace('93.', '') for s in file_list]
file_list = [s.replace('94.', '') for s in file_list]
file_list = [s.replace('95.', '') for s in file_list]
file_list = [s.replace('1.', '') for s in file_list]
file_list = [s.replace('2.', '') for s in file_list]
file_list = [s.replace('3.', '') for s in file_list]
file_list = [s.replace('4.', '') for s in file_list]
file_list = [s.replace('5.', '') for s in file_list]
file_list = [s.replace('6.', '') for s in file_list]
file_list = [s.replace('7.', '') for s in file_list]
file_list = [s.replace('8.', '') for s in file_list]
file_list = [s.replace('9.', '') for s in file_list]
file_list = [s.replace('OTU-Score0', 'OTU-S0') for s in file_list]
file_list = [s.replace('OTU-Score1', 'OTU-S1') for s in file_list]
file_list = [s.replace('OTU-Score2', 'OTU-S2') for s in file_list]
file_list = [s.replace('OTU-Score3', 'OTU-S3') for s in file_list]
file_list = [s.replace('Disease', 'DS') for s in file_list]
fig, axs = plt.subplots(n_model,6, figsize=(20, 20), facecolor='w', edgecolor='k')
fig.subplots_adjust(hspace =0.1, wspace=0.2)
ax = plt.gca()
ax.set_facecolor("white")
k=0
m=0
for i in range(0,n_model):
    ax.set_facecolor("white")
    for j in range(0,6):
        ax.set_facecolor("white")
        k=k+1
        v=EE.iloc[i,j]
        LL_value=[item for subitem in v for item in subitem]
        plt.subplot(n_model,6,k)
        if i==0:
            plt.title(file_list_response[j], fontsize=22)
            ax.set_facecolor("white")
        if j==0:
            plt.ylabel(file_list[i], rotation=0, fontsize=16,loc="center",multialignment='center', labelpad=60)
        ax=sns.boxplot(LL_value)
        ax.set_facecolor("white")
        plt.xlim(0,1)
        ax.axes.xaxis.set_ticklabels([])
        plt.axvline(0.75, color='black',ls = '--') 
        if i==n_model-1:
            ax.axes.xaxis.set_ticklabels([0,0.2,0.4,0.6,0.8,1])
fig.text(0.5, 0.1, 'Weighted F1-score', ha='center', va='center',fontsize=25)
plt.savefig(path_figure+'4-ALL-RF.png',bbox_inches = 'tight',dpi=600,facecolor='white')


# # 5-ALL-BNN

# In[31]:


file_list_response=[]
for file_response in os.listdir(path_BNN):
    if (file_response != '.DS_Store') & (file_response != 'Icon'): 
        print(file_response)
        file_list_response.append(file_response)  
        path_r= path_BNN+file_response
        file_list=[]
        os.chdir(path_r)
        for file in os.listdir(path_r):  
            if (file != '.DS_Store')& (file != 'Icon\r'): 
                if file=='classification_RF.xlsx': 
                    print(file)
                    os.rename('classification_RF.xlsx', '1.ALL-OTU.xlsx')
                elif file=='classification_RF_FS_notImportant.xlsx': 
                    print(file)
                    os.rename('classification_RF_FS_notImportant.xlsx', '2.OTU-Score0.xlsx')
                elif file=='classification_RF_FS_ML.xlsx': 
                    print(file)
                    os.rename('classification_RF_FS_ML.xlsx', '3.OTU-Score1.xlsx')

                elif file=='classification_RF_FS_Netcomi.xlsx': 
                    print(file)
                    os.rename('classification_RF_FS_Netcomi.xlsx', '4.OTU-Score2.xlsx')
                elif file=='classification_RF_FS_score3.xlsx': 
                    print(file)
                    os.rename('classification_RF_FS_score3.xlsx', '5.OTU-Score3.xlsx')

                    
                ###############################################    
                elif file=='classification_RF_alpha_diversity.xlsx': 
                    print(file)
                    os.rename('classification_RF_alpha_diversity.xlsx', '6.Alpha.xlsx')
                elif file=='classification_RF_soil_chemistry_2.xlsx': 
                    print(file)
                    os.rename('classification_RF_soil_chemistry_2.xlsx', '7.Soil.xlsx')
                elif file=='classification_RF_disease_suppression_2.xlsx': 
                    print(file)
                    os.rename('classification_RF_disease_suppression_2.xlsx', '8.Disease.xlsx')
                elif file=='classification_alpha+soil.xlsx': 
                    print(file)
                    os.rename('classification_alpha+soil.xlsx', '9.Alpha+Soil.xlsx')
                elif file=='classification_RF+soil+disease.xlsx': 
                    print(file)
                    os.rename('classification_RF+soil+disease.xlsx', '10.Soil+Disease.xlsx')
                elif file=='classification_alpha+soil+disease.xlsx': 
                    print(file)
                    os.rename('classification_alpha+soil+disease.xlsx', '11.Alpha+Soil+Disease.xlsx')
                    
                 ############################################### 
                 
                elif file=='classification_RF_FS_soil_chemistry.xlsx': 
                    print(file)
                    os.rename('classification_RF_FS_soil_chemistry.xlsx', '12.OTU-Score3+Soil.xlsx')
                elif file=='classification_RF_FS_disease_suppression.xlsx': 
                    print(file)
                    os.rename('classification_RF_FS_disease_suppression.xlsx', '13.OTU-Score3+Disease.xlsx')
                elif file=='classification_important_otu+soil+disease.xlsx': 
                    print(file)
                    os.rename('classification_important_otu+soil+disease.xlsx', '14.OTU-Score3+Soil+Disease.xlsx')                   
                    


# In[32]:


for file_response in os.listdir(path_BNN):
    if (file_response != '.DS_Store') & (file_response != 'Icon'): 
        print(file_response)
        file_list_response.append(file_response)  
        path_r= path_BNN+file_response
        file_list=[]
        os.chdir(path_r)
        for file in os.listdir(path_r):  
            if (file != '.DS_Store')& (file != 'Icon\r'): 
                if file=='10.Soil+Disease.xlsx': 
                    print(file)
                    os.rename('10.Soil+Disease.xlsx', '91.Soil+Disease.xlsx')
                elif file=='11.Alpha+Soil+Disease.xlsx': 
                    print(file)
                    os.rename('11.Alpha+Soil+Disease.xlsx', '92.Alpha+Soil+Disease.xlsx')
                elif file=='12.OTU-Score3+Soil.xlsx': 
                    print(file)
                    os.rename('12.OTU-Score3+Soil.xlsx', '93.OTU-Score3+Soil.xlsx')
                elif file=='13.OTU-Score3+Disease.xlsx': 
                    print(file)
                    os.rename('13.OTU-Score3+Disease.xlsx', '94.OTU-Score3+Disease.xlsx')
                elif file=='14.OTU-Score3+Soil+Disease.xlsx': 
                    print(file)
                    os.rename('14.OTU-Score3+Soil+Disease.xlsx', '95.OTU-Score3+Soil+Disease.xlsx')


# In[33]:


ResultFinal=pd.DataFrame()
file_list_response=[]
for file_response in sorted(os.listdir(path_BNN)):
    if (file_response != '.DS_Store') & (file_response != 'Icon'): 
        print(file_response)
        file_list_response.append(file_response)  
        path_r= path_BNN+file_response
        Result=pd.DataFrame()
        file_list=[]
        for file in sorted(os.listdir(path_r)):  
            if (file != '.DS_Store')& (file != 'Icon\r'):   
                file_list.append(file) 
                wb = openpyxl.load_workbook(path_BNN+file_response+'/'+file)
                sheet_list = wb.sheetnames
                ALL_value = []
                for sheet in range(1,len(sheet_list)+1):
                    response = pd.read_excel(path_BNN+file_response+'/'+file, sheet_name=sheet_list[sheet-1])
                    ALL_value.append(response.iloc[4,1:].values)
                CC= [file ,ALL_value]
        
                Result = pd.concat( [Result, pd.DataFrame(CC).transpose()],axis=0) 
        Result.columns=[file, file_response]
        print(Result.shape)      
        ResultFinal=  pd.concat([ResultFinal, pd.DataFrame(Result)],axis=1)


# In[34]:


n_model=ResultFinal.shape[0]
file_list_response= ['Yield_Plant','Yield_Meter', 'Black_Scurf','Scab','Scabpit','Scabsuper']
EE = ResultFinal[file_list_response]
file_list = [s.replace('.xlsx', '') for s in file_list]
file_list = [s.replace('91.', '') for s in file_list]
file_list = [s.replace('92.', '') for s in file_list]
file_list = [s.replace('93.', '') for s in file_list]
file_list = [s.replace('94.', '') for s in file_list]
file_list = [s.replace('95.', '') for s in file_list]
file_list = [s.replace('1.', '') for s in file_list]
file_list = [s.replace('2.', '') for s in file_list]
file_list = [s.replace('3.', '') for s in file_list]
file_list = [s.replace('4.', '') for s in file_list]
file_list = [s.replace('5.', '') for s in file_list]
file_list = [s.replace('6.', '') for s in file_list]
file_list = [s.replace('7.', '') for s in file_list]
file_list = [s.replace('8.', '') for s in file_list]
file_list = [s.replace('9.', '') for s in file_list]
file_list = [s.replace('OTU-Score0', 'OTU-S0') for s in file_list]
file_list = [s.replace('OTU-Score1', 'OTU-S1') for s in file_list]
file_list = [s.replace('OTU-Score2', 'OTU-S2') for s in file_list]
file_list = [s.replace('OTU-Score3', 'OTU-S3') for s in file_list]
file_list = [s.replace('Disease', 'DS') for s in file_list]


fig, axs = plt.subplots(n_model,6, figsize=(20, 20), facecolor='w', edgecolor='k')
fig.subplots_adjust(hspace =0.1, wspace=0.2)
ax = plt.gca()
ax.set_facecolor("white")
k=0
m=0
#fig, axxx = plt.subplots()
for i in range(0,n_model):
    ax.set_facecolor("white")
    for j in range(0,6):
        ax.set_facecolor("white")
        k=k+1
        v=EE.iloc[i,j]
        LL_value=[item for subitem in v for item in subitem]
        plt.subplot(n_model,6,k)
        if i==0:
            plt.title(file_list_response[j], fontsize=22)
            ax.set_facecolor("white")
        if j==0:
            plt.ylabel(file_list[i], rotation=0, fontsize=16,loc="center",multialignment='center', labelpad=60)
        ax=sns.boxplot(LL_value)
        ax.set_facecolor("white")
        plt.xlim(0,1)
        ax.axes.xaxis.set_ticklabels([])
        plt.axvline(0.75, color='black',ls = '--')    
        if i==n_model-1:
            ax.axes.xaxis.set_ticklabels([0,0.2,0.4,0.6,0.8,1])
fig.text(0.5, 0.1, 'Weighted F1-score', ha='center', va='center',fontsize=25)
plt.savefig(path_figure+'5-ALL-BNN.png',bbox_inches = 'tight',dpi=600,facecolor='white')


# # 6-Original-Aug-RF-BNN

# In[35]:


path_list = []

for root, dirs, files in os.walk(path_original, topdown=False):
    for path in dirs:
        path_list.append(path)
        
wb = openpyxl.load_workbook(path_original+path+'/1.ALL-OTU.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list)
    for folder in path_list:
        data_temp = pd.read_excel(path_original+'/'+folder+'/1.ALL-OTU.xlsx', sheet_name=sheet_name)
        temp_df[folder] = data_temp.iloc[4][1:].values

    results_dic[sheet_name] = temp_df

main_df = pd.DataFrame(columns=['Data', 'Response'])
for key in results_dic.keys():
    for col in results_dic[key]:
        temp_df = pd.DataFrame()
        temp_df['Data'] = results_dic[key][col].values
        temp_df['Response'] = col
        temp_df['Level'] = key
        main_df = pd.concat([main_df, temp_df]) 
main_df['Method']='Original Data'


# In[36]:


path_list = []

for root, dirs, files in os.walk(path_aug, topdown=False):
    for path in dirs:
        path_list.append(path)
        
wb = openpyxl.load_workbook(path_aug+path+'/classification_RF.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list)
    for folder in path_list:
        data_temp = pd.read_excel(path_aug+'/'+folder+'/classification_RF.xlsx', sheet_name=sheet_name)
        temp_df[folder] = data_temp.iloc[4][1:].values

    results_dic[sheet_name] = temp_df

main_df2 = pd.DataFrame(columns=['Data', 'Response'])
for key in results_dic.keys():
    for col in results_dic[key]:
        temp_df = pd.DataFrame()
        temp_df['Data'] = results_dic[key][col].values
        temp_df['Response'] = col
        temp_df['Level'] = key
        main_df2 = pd.concat([main_df2, temp_df]) 
        
main_df2['Method']='Augmented data'


# In[37]:


#BNN


# In[38]:


path_list = []

for root, dirs, files in os.walk(path_BNN, topdown=False):
    for path in dirs:
        path_list.append(path)
        
wb = openpyxl.load_workbook(path_BNN+path+'/1.ALL-OTU.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list)
    for folder in path_list:
        data_temp = pd.read_excel(path_BNN+'/'+folder+'/1.ALL-OTU.xlsx', sheet_name=sheet_name)
        temp_df[folder] = data_temp.iloc[4][1:].values

    results_dic[sheet_name] = temp_df

main_df_BN = pd.DataFrame(columns=['Data', 'Response'])
for key in results_dic.keys():
    for col in results_dic[key]:
        temp_df = pd.DataFrame()
        temp_df['Data'] = results_dic[key][col].values
        temp_df['Response'] = col
        temp_df['Level'] = key
        main_df_BN = pd.concat([main_df_BN, temp_df]) 
main_df_BN['Method']='Original Data'


# In[39]:


path_list = []

for root, dirs, files in os.walk(path_aug_BNN, topdown=False):
    for path in dirs:
        path_list.append(path)
        
wb = openpyxl.load_workbook(path_aug_BNN+path+'/classification_BNN.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list)
    for folder in path_list:
        data_temp = pd.read_excel(path_aug_BNN+'/'+folder+'/classification_BNN.xlsx', sheet_name=sheet_name)
        temp_df[folder] = data_temp.iloc[4][1:].values

    results_dic[sheet_name] = temp_df

main_df_BNN2 = pd.DataFrame(columns=['Data', 'Response'])
for key in results_dic.keys():
    for col in results_dic[key]:
        temp_df = pd.DataFrame()
        temp_df['Data'] = results_dic[key][col].values
        temp_df['Response'] = col
        temp_df['Level'] = key
        main_df_BNN2 = pd.concat([main_df_BNN2, temp_df]) 
        
main_df_BNN2['Method']='Augmented data'


# In[40]:


fig, axs = plt.subplots(2,2, figsize=(20,9), facecolor='w', edgecolor='k')

plt.subplot(2,2,1)
AA = sns.boxplot(x="Response", hue="Level", y="Data", data=main_df,palette='Set2',hue_order =hue_order ,order=order)
plt.ylabel('',fontsize=16)
plt.title('Original Data',fontsize=20)
plt.xticks(fontsize=14)
plt.yticks(fontsize=15)
AA.grid(axis= 'y')
plt.legend([],[], frameon=False)
plt.xlabel('')
plt.yticks([0.2,0.4,0.6,0.8,1])
plt.ylim(0,1)

ax2=plt.subplot(2,2,2)
AA2 = sns.boxplot(x="Response", hue="Level", y="Data", data=main_df2,palette='Set2',hue_order =hue_order,order=order )
plt.ylabel('')
plt.xlabel('')
plt.title('Augmented Data',fontsize=20)
plt.xticks(fontsize=14)
plt.legend([],[], frameon=False)
ax = axs[0]
plt.ylim(0,1)
AA2.axes.yaxis.set_ticklabels([])
AA2.grid(axis= 'y')
fig.subplots_adjust(hspace =0, wspace=0)
plt.ylabel('RF',fontsize=20)
ax2.yaxis.set_label_position("right")

plt.subplot(2,2,3)
AA = sns.boxplot(x="Response", hue="Level", y="Data", data=main_df_BN,palette='Set2',hue_order =hue_order ,order=order)
plt.xticks(fontsize=14)
plt.yticks(fontsize=15)
AA.grid(axis= 'y')
plt.legend([],[], frameon=False)
plt.xlabel('')
plt.ylabel('')
plt.yticks([0,0.2,0.4,0.6,0.8])
plt.ylim(0,1)

ax=plt.subplot(2,2,4)
AA2 = sns.boxplot(x="Response", hue="Level", y="Data", data=main_df_BNN2,palette='Set2',hue_order =hue_order,order=order )
plt.ylabel('')
plt.xlabel('')
plt.xticks(fontsize=14)
AA2.axes.yaxis.set_ticklabels([])
AA2.grid(axis= 'y')
plt.ylim(0,1)
legend = plt.legend(loc='lower center',fontsize=13,mode = "expand", ncol = 5)
fig.text(0.1, 0.5, 'Weighted F1-score', ha='center', va='center',fontsize=20, rotation='vertical')
plt.ylabel('BNN',fontsize=20)
ax.yaxis.set_label_position("right")
plt.savefig(path_figure+'6-Original-Aug-RF-BNN.png',bbox_inches = 'tight',dpi=600,facecolor='white')


# # 7.CompareToRandom

# In[41]:


R1=pd.read_excel(data_path2+'/Scabpit/RMethod1.xlsx', sheet_name='Phylum')
R2=pd.read_excel(data_path2+'/Scabpit/RMethod2.xlsx', sheet_name='Phylum')
R3=pd.read_excel(data_path2+'/Scabpit/RMethod3.xlsx', sheet_name='Phylum')
R4=pd.read_excel(data_path2+'/Scabpit/RMethod4.xlsx', sheet_name='Phylum')


# In[42]:


R11=pd.read_excel(data_path2+'/Scabsuper/RMethod1.xlsx', sheet_name='Phylum')
R22=pd.read_excel(data_path2+'/Scabsuper/RMethod2.xlsx', sheet_name='Phylum')
R33=pd.read_excel(data_path2+'/Scabsuper/RMethod3.xlsx', sheet_name='Phylum')
R44=pd.read_excel(data_path2+'/Scabsuper/RMethod4.xlsx', sheet_name='Phylum')


# In[43]:


fig, ax = plt.subplots(2,4, figsize=(15, 8), facecolor='w', edgecolor='k')
fig.subplots_adjust(hspace =0, wspace=0)
a=0.92
b=0.888888889
plt.subplot(2,4,1)
r1=R1
plt.scatter(r1['0'],r1['1'])
plt.scatter(a,b,s=100,c='red')
plt.xlabel('')
plt.xlim(0,1)
plt.ylim(0,1)
plt.yticks([0.2,0.4,0.6,0.8,1])
plt.xticks([])
plt.title('Strategy 1',fontsize=18)

plt.subplot(2,4,2)
r2=R2
plt.scatter(r2['0'],r2['1'])
plt.scatter(a,b,s=100,c='red')
plt.xlabel('')
plt.ylabel('')
plt.xlim(0,1)
plt.ylim(0,1)
plt.yticks([])
plt.xticks([])
plt.title('Strategy 2',fontsize=18)

plt.subplot(2,4,3)
r3=R3
plt.scatter(r3['0'],r3['1'])
plt.scatter(a,b,s=100,c='red')
plt.xlabel('')
plt.ylabel('')
plt.xlim(0,1)

plt.ylim(0,1)
plt.yticks([])
plt.xticks([])
plt.title('Strategy 3',fontsize=18)

ax = plt.subplot(2,4,4)
r4=R4
plt.scatter(r4['0'],r4['1'])
plt.scatter(a,b,s=100,c='red')
plt.xlabel('')
plt.title('Strategy 4',fontsize=18)
plt.ylabel('Scabpit',fontsize=18)
ax.yaxis.set_label_position("right")
ax.yaxis.tick_right()
# plt.ylabel('')
plt.xlim(0,1)
plt.ylim(0,1)
plt.xticks([])
plt.yticks([])
##############################
a=0.307692308
b=0.7
ax = plt.subplot(2,4,5)
r1=R11
plt.scatter(r1['0'],r1['1'])
plt.scatter(a,b,s=100,c='red')
plt.xlabel('')
plt.xlim(0,1)
plt.ylim(0,1)
plt.xticks([0.2,0.4,0.6,0.8])
plt.yticks([0,0.2,0.4,0.6,0.8])
plt.subplot(2,4,6)
r2=R22
plt.scatter(r2['0'],r2['1'])
plt.scatter(a,b,s=100,c='red')
plt.xlabel('')
plt.ylabel('')
plt.xlim(0,1)
plt.ylim(0,1)
plt.yticks([])
plt.xticks([0.2,0.4,0.6,0.8])

plt.subplot(2,4,7)
r3=R33
plt.scatter(r3['0'],r3['1'])
plt.scatter(a,b,s=100,c='red')
plt.xlabel('')
plt.ylabel('')
plt.xlim(0,1)

plt.ylim(0,1)
plt.yticks([])
plt.xticks([0.2,0.4,0.6,0.8])

ax = plt.subplot(2,4,8)
r4=R44
plt.scatter(r4['0'],r4['1'])
plt.scatter(a,b,s=100,c='red')
plt.xlabel('')
plt.ylabel('Scabsuper',fontsize=18)
ax.yaxis.set_label_position("right")
ax.yaxis.tick_right()
plt.xlim(0,1)
plt.ylim(0,1)
plt.xticks([0.2,0.4,0.6,0.8,1])
plt.yticks([])
fig.text(0.5, 0.06, 'F1-score Label 0', ha='center', va='center',fontsize=20)
fig.text(0.09, 0.5, 'F1-score Label 1', ha='center', va='center',fontsize=20, rotation='vertical')
plt.savefig(path_figure+'7-CompareToRandom.png',bbox_inches = 'tight',dpi=600,facecolor='white')


# # 8-EVRF

# In[44]:


path_list = []

for root, dirs, files in os.walk(data_path2, topdown=False):
    for path in dirs:
        path_list.append(path)
        
wb = openpyxl.load_workbook(data_path2+'/'+path+'/EV_Value.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list)
    for folder in path_list:
        data_temp = pd.read_excel(data_path2+'/'+folder+'/EV_Value.xlsx', sheet_name=sheet_name)
        temp_df[folder] = data_temp.iloc[:,1].values

    results_dic[sheet_name] = temp_df
main_df = pd.DataFrame(columns=['Data', 'Response'])
for key in results_dic.keys():
    for col in results_dic[key]:
        temp_df = pd.DataFrame()
        temp_df['Data'] = results_dic[key][col].values
        temp_df['Response'] = col
        temp_df['Level'] = key
        main_df = pd.concat([main_df, temp_df])


# In[45]:


path_list = []
for root, dirs, files in os.walk(path_aug, topdown=False):
    for path in dirs:
        path_list.append(path)
        
wb = openpyxl.load_workbook(path_aug+'/'+path+'/EV_Value.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list)
    for folder in path_list:
        data_temp = pd.read_excel(path_aug+'/'+folder+'/EV_Value.xlsx', sheet_name=sheet_name)
        temp_df[folder] = data_temp.iloc[:,1].values

    results_dic[sheet_name] = temp_df
main_df2 = pd.DataFrame(columns=['Data', 'Response'])
for key in results_dic.keys():
    for col in results_dic[key]:
        temp_df = pd.DataFrame()
        temp_df['Data'] = results_dic[key][col].values
        temp_df['Response'] = col
        temp_df['Level'] = key
        main_df2 = pd.concat([main_df2, temp_df])


# In[47]:


fig, axs = plt.subplots(1,2, figsize=(20,5), facecolor='w', edgecolor='k')
plt.subplot(1,2,1)
order= ['Yield_Plant','Yield_Meter', 'Black_Scurf','Scab','Scabpit','Scabsuper']
AA = sns.boxplot(x="Response", hue="Level", y="Data", data=main_df,palette='Set2',hue_order =hue_order ,order=order)
plt.axhline(0.05, color='r',ls = '--') # vertical
plt.ylabel('EV Value',fontsize=16)
plt.xticks(fontsize=13)
plt.xlabel('')
plt.yticks(fontsize=13)
AA.grid(axis= 'y')
plt.legend([],[], frameon=False)
plt.ylim(-0.01,1.17)
plt.yticks([0,0.2,0.4,0.6,0.8,1])
plt.title('Original Data',fontsize=18)
plt.subplot(1,2,2)
order= ['Yield_Plant','Yield_Meter', 'Black_Scurf','Scab','Scabpit','Scabsuper']
AA2 = sns.boxplot(x="Response", hue="Level", y="Data", data=main_df2,palette='Set2',hue_order =hue_order,order=order )
plt.axhline(0.05, color='r',ls = '--') # vertical
plt.ylabel('')
plt.title('Augmented Data',fontsize=18)
plt.xlabel('')
plt.xticks(fontsize=15)
ax = axs[0]
plt.ylim(-0.01,1.17)
legend = plt.legend(loc='upper center',fontsize=13,mode = "expand", ncol = 5)
AA2.axes.yaxis.set_ticklabels([])
AA2.grid(axis= 'y')
fig.subplots_adjust(hspace =0, wspace=0)
plt.savefig(path_figure+'8-EVRF.png',bbox_inches = 'tight',dpi=600,facecolor='white')


# # 9-RF-BNN-4Dis

# In[28]:


path_list = []

for root, dirs, files in os.walk(path_Disease, topdown=False):
    for path in dirs:
        path_list.append(path)
        
wb = openpyxl.load_workbook(path_Disease+'/'+path+'/1.ALL-OTU.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list)
    for folder in path_list:
        data_temp = pd.read_excel(path_Disease+'/'+folder+'/1.ALL-OTU.xlsx', sheet_name=sheet_name)
        temp_df[folder] = data_temp.iloc[4][1:].values

    results_dic[sheet_name] = temp_df


# In[29]:


main_df = pd.DataFrame(columns=['Data', 'Response'])
for key in results_dic.keys():
    for col in results_dic[key]:
        temp_df = pd.DataFrame()
        temp_df['Data'] = results_dic[key][col].values
        temp_df['Response'] = col
        temp_df['Level'] = key
        main_df = pd.concat([main_df, temp_df]) 


# BNN

# In[30]:


path_list = []

for root, dirs, files in os.walk(path_Disease_BNN, topdown=False):
    for path in dirs:
        path_list.append(path)
        
wb = openpyxl.load_workbook(path_Disease_BNN+'/'+path+'/1.ALL-OTU.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list)
    for folder in path_list:
        data_temp = pd.read_excel(path_Disease_BNN+'/'+folder+'/1.ALL-OTU.xlsx', sheet_name=sheet_name)
        temp_df[folder] = data_temp.iloc[4][1:].values

    results_dic[sheet_name] = temp_df


# In[31]:


main_df2 = pd.DataFrame(columns=['Data', 'Response'])
for key in results_dic.keys():
    for col in results_dic[key]:
        temp_df = pd.DataFrame()
        temp_df['Data'] = results_dic[key][col].values
        temp_df['Response'] = col
        temp_df['Level'] = key
        main_df2 = pd.concat([main_df2, temp_df]) 


# In[32]:


fig, axs = plt.subplots(1,2, figsize=(20,5), facecolor='w', edgecolor='k')

plt.subplot(1,2,1)
order= [ 'Black_Scurf','Scab','Scabpit','Scabsuper']
AA = sns.boxplot(x="Response", hue="Level", y="Data", data=main_df,palette='Set2',hue_order =hue_order ,order=order)
plt.xticks(fontsize=18)
plt.yticks(fontsize=15)
AA.grid(axis= 'y')
plt.legend([],[], frameon=False)
plt.ylabel('Weighted F1-score',fontsize=16)
plt.xlabel('')
plt.title('Random Forest',fontsize=20)
plt.ylim(0,1)

plt.subplot(1,2,2)
order= ['Black_Scurf','Scab','Scabpit','Scabsuper']
AA2 = sns.boxplot(x="Response", hue="Level", y="Data", data=main_df2,palette='Set2',hue_order =hue_order,order=order )
#plt.xticks(fontweight='bold')
plt.ylabel('')
plt.xlabel('')
plt.xticks(fontsize=18)
plt.title('Bayesian NN',fontsize=20)
ax = axs[0]

plt.ylim(0,1)

legend = plt.legend(loc='lower right',fontsize=13,mode = "expand", ncol = 5)

AA2.axes.yaxis.set_ticklabels([])
AA2.grid(axis= 'y')
fig.subplots_adjust(hspace =0, wspace=0)

plt.savefig(path_figure+'9-RF-BNN-4Dis.png',bbox_inches = 'tight',dpi=600,facecolor='white')


# # 10-RF-BNN-Normalized-Scabpit

# In[33]:


Result=pd.DataFrame()
file_list_response=[]
for file_response in os.listdir(path_Disease):
    if (file_response != '.DS_Store') & (file_response != 'Icon')& (file_response == 'Scabpit'): 
        print(file_response)
        file_list_response.append(file_response)  
        path_r= path_Disease+file_response
        file_list=[]
        wb = openpyxl.load_workbook(path_Disease+file_response+'/'+'1.ALL-OTU.xlsx')
        sheet_list = wb.sheetnames
        k=0
        for sheet in range(1,len(sheet_list)+1):
            #print(sheet-1)
            F1 = pd.read_excel(path_Disease+file_response+'/'+'1.ALL-OTU.xlsx', sheet_name=sheet_list[sheet-1])
            for nc in range(1,21):
                CC= [file_response ,sheet_list[sheet-1],nc, F1.iloc[4,nc]]
                Result = pd.concat( [Result, pd.DataFrame(CC).transpose()],axis=0)
                k=k+1
Result.columns=['response', 'level','Normalization method','wF1-score']


# In[34]:


Method=['TSS+none',
'TSS+pseudo',
'TSS+multRepl',
'TSS+bayesMult',
'CSS+none',
'CSS+pseudo',
'CSS+multRepl',
'CSS+bayesMult',
'COM+none',
'COM+pseudo',
'COM+multRepl',
'COM+bayesMult',
'rarefy+none',
'rarefy+pseudo',
'rarefy+multRepl',
'rarefy+bayesMult',
'clr+none',
'clr+pseudo',
'clr+multRepl',
'clr+bayesMult']


# BNN

# In[35]:


Result2=pd.DataFrame()
file_list_response=[]
for file_response in os.listdir(path_Disease_BNN):
    if (file_response != '.DS_Store') & (file_response != 'Icon')& (file_response == 'Scabpit'): 
        print(file_response)
        file_list_response.append(file_response)  
        path_r= path_Disease_BNN+file_response
        file_list=[]
        wb = openpyxl.load_workbook(path_Disease_BNN+file_response+'/'+'1.ALL-OTU.xlsx')
        sheet_list = wb.sheetnames
        k=0
        for sheet in range(1,len(sheet_list)+1):
            F1 = pd.read_excel(path_Disease_BNN+file_response+'/'+'1.ALL-OTU.xlsx', sheet_name=sheet_list[sheet-1])
            for nc in range(1,21):
                CC= [file_response ,sheet_list[sheet-1],nc, F1.iloc[4,nc]]
                Result2 = pd.concat( [Result2, pd.DataFrame(CC).transpose()],axis=0)
                k=k+1
Result2.columns=['response', 'level','Normalization method','wF1-score']


# In[36]:


Result['class']='Random Forest'
Result2['class']='Bayesian NN'


# In[37]:


RRR=pd.concat([Result,Result2],axis=0)


# In[38]:


plt.figure(figsize=(15,4), dpi=600)
g = sns.catplot(
    data=RRR, x='Normalization method', y='wF1-score', hue='level', col="class",palette='Set2',s=10,hue_order = hue_order,
    height=4, aspect=2)


g.set_axis_labels("", 'Weighted F1-score',fontsize=16)
g.set_xticklabels(Method,size= 16)
g.set_titles("{col_name}",size= 18)
g.set(ylim=(0.6, 1))
g.set_xticklabels(rotation =90)
g.despine(left=True)
g._legend.remove()
plt.legend(loc='lower right',fontsize=15,mode = "expand", ncol = 5)

plt.savefig(path_figure+'10-RF-BNN-Normalized-Scabpit.png',bbox_inches = 'tight',dpi=600,facecolor='white')


# # 11-RF-BNN-SelectedFeatures_Scabpit

# In[39]:


ResultFinal=pd.DataFrame()
file_list_response=[]
for file_response in os.listdir(path_Disease):
    if (file_response != '.DS_Store') & (file_response != 'Icon')& (file_response == 'Scabpit'): 
        print(file_response)
        file_list_response.append(file_response)  
        path_r= path_Disease+file_response
        Result=pd.DataFrame()
        file_list=[]
        for file in sorted(os.listdir(path_r))[0:6]:  
            if (file != '.DS_Store')& (file != 'Icon\r'):   
                file_list.append(file) 
                wb = openpyxl.load_workbook(path_Disease+file_response+'/'+file)
                sheet_list = wb.sheetnames
                ALL_value = []
                for sheet in range(1,len(sheet_list)+1):
                    response = pd.read_excel(path_Disease+file_response+'/'+file, sheet_name=sheet_list[sheet-1])
                    CC= [file ,sheet_list[sheet-1],round(response.iloc[4,1], 2)]
                    print(CC)
                    Result = pd.concat( [Result, pd.DataFrame(CC).transpose()],axis=0) 
        Result.columns=['Methods', 'level', 'Weighted F1-score']
        print(Result.shape)    
        ResultFinal=  pd.concat([ResultFinal, pd.DataFrame(Result)],axis=1)


# BNN

# In[40]:


ResultFinal2=pd.DataFrame()
file_list_response=[]
for file_response in os.listdir(path_Disease_BNN):
    if (file_response != '.DS_Store') & (file_response != 'Icon')& (file_response == 'Scabpit'):  
        print(file_response)
        file_list_response.append(file_response)  
        path_r= path_Disease_BNN+file_response
        Result=pd.DataFrame()
        file_list=[]
        for file in sorted(os.listdir(path_r))[0:5]:  
            if (file != '.DS_Store')& (file != 'Icon\r'):   
                file_list.append(file) 
                wb = openpyxl.load_workbook(path_Disease_BNN+file_response+'/'+file)
                sheet_list = wb.sheetnames
                ALL_value = []
                for sheet in range(1,len(sheet_list)+1):
                    response = pd.read_excel(path_Disease_BNN+file_response+'/'+file, sheet_name=sheet_list[sheet-1])
                    CC= [file ,sheet_list[sheet-1],round(response.iloc[4,1], 2)]
                    print(CC)
                    Result = pd.concat( [Result, pd.DataFrame(CC).transpose()],axis=0) 
        Result.columns=['Methods', 'level', 'Weighted F1-score']
        print(Result.shape)     
        ResultFinal2=  pd.concat([ResultFinal2, pd.DataFrame(Result)],axis=1)


# In[41]:


ResultFinal['class']='Random Forest'
ResultFinal2['class']='Bayesian NN'


# In[42]:


RRR=pd.concat([ResultFinal,ResultFinal2],axis=0)


# In[43]:


plt.figure(figsize=(15,4), dpi=600)
g = sns.catplot(
    data=RRR, x='Methods', y='Weighted F1-score', hue='level', col="class",palette='Set2',kind='bar',hue_order = hue_order,
    height=4, aspect=2)

g.set_axis_labels("", 'Weighted F1-score',fontsize=16)
g.set_xticklabels(['ALL-OTU','OTU-S0','OTU-S1','OTU-S2','OTU-S3'],size= 16)
g.set_titles("{col_name}",size= 18)
g.set(ylim=(0.6, 1))
#g.set_xticklabels(rotation =90)
g.despine(left=True)
g._legend.remove()
plt.legend(loc='lower right',fontsize=15,mode = "expand", ncol = 5)

plt.savefig(path_figure+'11-RF-BNN-SelectedFeatures_Scabpit.png',bbox_inches = 'tight',dpi=600,facecolor='white')


# # 12-DT-RF-Scabpit

# based on 20 generated datasets

# In[44]:


Result=pd.DataFrame()
file_list_response=[]
for file_response in os.listdir(path_original):
    if (file_response != '.DS_Store') & (file_response != 'Icon'): 
        print(file_response)
        file_list_response.append(path_original)  
        path_r= path_original+file_response
        file_list=[]
        wb = openpyxl.load_workbook(path_original+file_response+'/'+'1.ALL-OTU.xlsx')
        sheet_list = wb.sheetnames
        k=0
        for sheet in range(1,len(sheet_list)+1):
            #print(sheet-1)
            F1 = pd.read_excel(path_original+file_response+'/'+'1.ALL-OTU.xlsx', sheet_name=sheet_list[sheet-1])
            for nc in range(1,21):
                CC= [file_response ,sheet_list[sheet-1],nc, F1.iloc[4,nc]]
                Result = pd.concat( [Result, pd.DataFrame(CC).transpose()],axis=0)
                k=k+1
Result.columns=['response', 'L','NM','wF1-score']
Result['Aug']=0


# In[45]:


Result2=pd.DataFrame()
file_list_response=[]
for file_response in os.listdir(path_aug):
    if (file_response != '.DS_Store') & (file_response != 'Icon'): 
        print(file_response)
        file_list_response.append(file_response)  
        path_r= path_aug+file_response
        file_list=[]
        wb = openpyxl.load_workbook(path_aug+file_response+'/'+'classification_RF.xlsx')
        sheet_list = wb.sheetnames
        k=0
        for sheet in range(1,len(sheet_list)+1):
            #print(sheet-1)
            F1 = pd.read_excel(path_aug+file_response+'/'+'classification_RF.xlsx', sheet_name=sheet_list[sheet-1])
            for nc in range(1,21):
                CC= [file_response ,sheet_list[sheet-1],nc, F1.iloc[4,nc]]
                Result2 = pd.concat( [Result2, pd.DataFrame(CC).transpose()],axis=0)
                k=k+1
        #ALL_value=[item for subitem in ALL_value for item in subitem]
            #ALL_value.append(CC)
Result2.columns=['response', 'L','NM','wF1-score']
Result2['Aug']=1


# In[46]:


REE=pd.concat([Result,Result2],axis=0)
#For Scabpit
tableScab=REE[REE['response']=='Scabpit']


# In[47]:


tableScab_2=tableScab[['L','NM','Aug','wF1-score']]
data_x = tableScab_2.drop('wF1-score', axis=1)
data_y = tableScab_2['wF1-score']
data_x_encoded = pd.get_dummies(data_x, drop_first=True)
dtree = DecisionTreeRegressor(max_depth=4)
dtree.fit(data_x_encoded , data_y)


# In[48]:


fig = plt.figure(figsize=((20,10)))
dot_data = tree.export_graphviz(dtree, out_file=None,impurity=False, 
                                feature_names=data_x_encoded.columns, label='root', 
                                filled=True,leaves_parallel=True, proportion=True,rotate=False,rounded=True,special_characters=True,precision=3)  

# Draw graph
fig = graphviz.Source(dot_data, format="png") 
fig.render(path_figure+'12-DT-RF-Scabpit',view = True)


# BN

# In[48]:


Result=pd.DataFrame()
file_list_response=[]
for file_response in os.listdir(path_BNN):
    if (file_response != '.DS_Store') & (file_response != 'Icon'): 
        print(file_response)
        file_list_response.append(path_BNN)  
        path_r= path_original+file_response
        file_list=[]
        wb = openpyxl.load_workbook(path_BNN+file_response+'/'+'1.ALL-OTU.xlsx')
        sheet_list = wb.sheetnames
        k=0
        for sheet in range(1,len(sheet_list)+1):
            #print(sheet-1)
            F1 = pd.read_excel(path_BNN+file_response+'/'+'1.ALL-OTU.xlsx', sheet_name=sheet_list[sheet-1])
            for nc in range(1,21):
                CC= [file_response ,sheet_list[sheet-1],nc, F1.iloc[4,nc]]
                Result = pd.concat( [Result, pd.DataFrame(CC).transpose()],axis=0)
                k=k+1
Result.columns=['response', 'L','NM','wF1-score']
Result['Aug']=0


# In[49]:


Result2=pd.DataFrame()
file_list_response=[]
for file_response in os.listdir(path_aug_BNN):
    if (file_response != '.DS_Store') & (file_response != 'Icon'): 
        print(file_response)
        file_list_response.append(file_response)  
        path_r= path_aug_BNN+file_response
        file_list=[]
        wb = openpyxl.load_workbook(path_aug_BNN+file_response+'/'+'classification_BNN.xlsx')
        sheet_list = wb.sheetnames
        k=0
        for sheet in range(1,len(sheet_list)+1):
            #print(sheet-1)
            F1 = pd.read_excel(path_aug_BNN+file_response+'/'+'classification_BNN.xlsx', sheet_name=sheet_list[sheet-1])
            for nc in range(1,21):
                CC= [file_response ,sheet_list[sheet-1],nc, F1.iloc[4,nc]]
                Result2 = pd.concat( [Result2, pd.DataFrame(CC).transpose()],axis=0)
                k=k+1
Result2.columns=['response', 'L','NM','wF1-score']
Result2['Aug']=1


# In[57]:


REE[REE['response']=='Scabpit']


# In[59]:


REE=pd.concat([Result,Result2],axis=0)
#For Scabpit
tableScab=REE[REE['response']=='Scabpit']


# In[60]:


REE.to_csv('/Users/rosa/Desktop/file1.csv')


# In[51]:


tableScab_2=tableScab[['L','NM','Aug','wF1-score']]
data_x = tableScab_2.drop('wF1-score', axis=1)
data_y = tableScab_2['wF1-score']
data_x_encoded = pd.get_dummies(data_x, drop_first=True)
dtree = DecisionTreeRegressor(max_depth=4)
dtree.fit(data_x_encoded , data_y)


# In[52]:


fig = plt.figure(figsize=((20,10)))
dot_data = tree.export_graphviz(dtree, out_file=None,impurity=False, 
                                feature_names=data_x_encoded.columns, label='root', 
                                filled=True,leaves_parallel=True, proportion=True,rotate=False,rounded=True,special_characters=True,precision=3)  


# In[53]:


dot_data
#0.5replaceby0
#&le;replaceby=


# In[54]:


dot_data='digraph Tree {\nnode [shape=box, style="filled, rounded", color="black", fontname="helvetica"] ;\ngraph [ranksep=equally, splines=polyline] ;\nedge [fontname="helvetica"] ;\n0 [label=<Aug = 0<br/>samples = 100.0%<br/>value = 0.775>, fillcolor="#eb9c64"] ;\n1 [label=<NM_13 = 0<br/>66.7%<br/>0.737>, fillcolor="#eca571"] ;\n0 -> 1 [labeldistance=2.5, labelangle=45, headlabel="True"] ;\n2 [label=<NM_3 = 0<br/>63.3%<br/>0.747>, fillcolor="#eca26e"] ;\n1 -> 2 ;\n3 [label=<NM_14 = 0<br/>60.0%<br/>0.758>, fillcolor="#eba06a"] ;\n2 -> 3 ;\n4 [label=<56.7%<br/>0.77>, fillcolor="#eb9d66"] ;\n3 -> 4 ;\n5 [label=<3.3%<br/>057>, fillcolor="#f5cdb1"] ;\n3 -> 5 ;\n6 [label=<L_Order = 0<br/>3.3%<br/>051>, fillcolor="#f5cfb3"] ;\n2 -> 6 ;\n7 [label=<2.5%<br/>0.623>, fillcolor="#f2be9a"] ;\n6 -> 7 ;\n8 [label=<0.8%<br/>0.336>, fillcolor="#ffffff"] ;\n6 -> 8 ;\n9 [label=<L_Order = 0<br/>3.3%<br/>046>, fillcolor="#f5d0b5"] ;\n1 -> 9 ;\n10 [label=<L_Family = 0<br/>2.5%<br/>0.615>, fillcolor="#f2c09c"] ;\n9 -> 10 ;\n11 [label=<1.7%<br/>0.755>, fillcolor="#eca16b"] ;\n10 -> 11 ;\n12 [label=<0.8%<br/>0.336>, fillcolor="#ffffff"] ;\n10 -> 12 ;\n13 [label=<0.8%<br/>0.336>, fillcolor="#ffffff"] ;\n9 -> 13 ;\n14 [label=<L_Phylum = 0<br/>33.3%<br/>0.852>, fillcolor="#e78b49"] ;\n0 -> 14 [labeldistance=2.5, labelangle=-45, headlabel="False"] ;\n15 [label=<NM_10 = 0<br/>16.7%<br/>0.884>, fillcolor="#e6843d"] ;\n14 -> 15 ;\n16 [label=<NM_18 = 0<br/>15.8%<br/>0.89>, fillcolor="#e5823b"] ;\n15 -> 16 ;\n17 [label=<15.0%<br/>0.896>, fillcolor="#e58139"] ;\n16 -> 17 ;\n18 [label=<0.8%<br/>0.785>, fillcolor="#ea9a60"] ;\n16 -> 18 ;\n19 [label=<0.8%<br/>0.765>, fillcolor="#eb9f67"] ;\n15 -> 19 ;\n20 [label=<NM_18 = 0<br/>16.7%<br/>0.819>, fillcolor="#e99254"] ;\n14 -> 20 ;\n21 [label=<NM_6 = 0<br/>15.8%<br/>0.828>, fillcolor="#e89051"] ;\n20 -> 21 ;\n22 [label=<15.0%<br/>0.834>, fillcolor="#e88f4f"] ;\n21 -> 22 ;\n23 [label=<0.8%<br/>0.731>, fillcolor="#eda673"] ;\n21 -> 23 ;\n24 [label=<0.8%<br/>0.652>, fillcolor="#f0b88f"] ;\n20 -> 24 ;\n{rank=same ; 0} ;\n{rank=same ; 1; 14} ;\n{rank=same ; 2; 9; 15; 20} ;\n{rank=same ; 3; 6; 10; 16; 21} ;\n{rank=same ; 4; 5; 7; 8; 11; 12; 13; 17; 18; 19; 22; 23; 24} ;\n}'


# In[55]:


# Draw graph
fig = graphviz.Source(dot_data, format="png") 
fig.render(path_figure+'13-DT-BNN-Scabpit',view = True)


# # 14-RF-BNN-Env-Scabpit

# In[60]:


ResultFinal=pd.DataFrame()
file_list_response=[]
for file_response in sorted(os.listdir(path_Disease)):
    if (file_response != '.DS_Store') & (file_response != 'Icon')&(file_response == 'Scabpit'): 
        print(file_response)
        file_list_response.append(file_response)  
        path_r= path_Disease+file_response
        Result=pd.DataFrame()
        file_list=[]
        for file in sorted(os.listdir(path_r)):  
            if (file != '.DS_Store')& (file != 'Icon\r') & (file[0] == '7') or (file[0] == '8')or (file[0:2] == '91') :  
                file_list.append(file) 
                wb = openpyxl.load_workbook(path_Disease+file_response+'/'+file)
                response = pd.read_excel(path_Disease+file_response+'/'+file)
                CC= [file ,response.iloc[4,1:].values]
        
                Result = pd.concat( [Result, pd.DataFrame(CC).transpose()],axis=0) 
        Result.columns=["response","Model"]
        print(Result.shape)      
        ResultFinal=  pd.concat([ResultFinal, pd.DataFrame(Result)],axis=1)


# In[61]:


data_1 = ResultFinal.iloc[0,1]
data_2 = ResultFinal.iloc[1,1]
data_3 = ResultFinal.iloc[2,1]
data = [data_1, data_2, data_3]
df= pd.DataFrame(data).T
df = df.rename(columns={k: f'Method{k+1}' for k in range(len(data))}).reset_index()
dataa=pd.melt(df,id_vars=['index'])
dataa['alg']='RF'


# BNN

# In[62]:


ResultFinal=pd.DataFrame()
file_list_response=[]
for file_response in sorted(os.listdir(path_Disease_BNN)):
    if (file_response != '.DS_Store') & (file_response != 'Icon')&(file_response == 'Scabpit'):  
        print(file_response)
        file_list_response.append(file_response)  
        path_r= path_Disease_BNN+file_response
        Result=pd.DataFrame()
        file_list=[]
        for file in sorted(os.listdir(path_r)):  
            if (file != '.DS_Store')& (file != 'Icon\r')& (file[0] == '7') or (file[0] == '8')or (file[0:2] == '91'):    
                file_list.append(file) 
                wb = openpyxl.load_workbook(path_Disease_BNN+file_response+'/'+file)
                response = pd.read_excel(path_Disease_BNN+file_response+'/'+file)
                CC= [file ,response.iloc[4,1:].values]
        
                Result = pd.concat( [Result, pd.DataFrame(CC).transpose()],axis=0) 
        Result.columns=["response","Model"]
        print(Result.shape)
                #Result.rename(columns={1:file_response}, inplace=True)      
        ResultFinal=  pd.concat([ResultFinal, pd.DataFrame(Result)],axis=1)
    
                
                #response = pd.read_excel(path_response+file_response+'/'+file, sheet_name="Order") 


# In[63]:


data_1 = ResultFinal.iloc[0,1]
data_2 = ResultFinal.iloc[1,1]
data_3 = ResultFinal.iloc[2,1]

data = [data_1, data_2, data_3]
df= pd.DataFrame(data).T
df = df.rename(columns={k: f'Method{k+1}' for k in range(len(data))}).reset_index()
dataa2=pd.melt(df,id_vars=['index'])
dataa2['alg']='BNN'
D=pd.concat([dataa,dataa2],axis=0)


# In[64]:


plt.figure(figsize=(9,4), dpi=600)
ax = plt.gca()
ax.set_ylim([0.3, 1])
#sns.boxplot(data = dataa, x='variable', y='value', hue='Level', palette='Set1')
AA=sns.boxplot(data = D, x='variable', y='value', hue='alg', palette="Set2")
AA.grid(axis= 'y')

#plt.xticks(fontsize=20)
plt.ylabel('Weighted F1-score',fontsize=16)
plt.xlabel('')
plt.yticks(fontsize=14)

ax.set_xticklabels(["Soil",'DS',"Soil+DS"],fontsize=18)
legend = plt.legend(loc='lower center',fontsize=17,mode = "expand", ncol = 2)
ax.set_facecolor("white")
plt.savefig(path_figure+'14-RF-BNN-Env-Scabpit.png',bbox_inches = 'tight',dpi=600,facecolor='white')


# # 15-RF-BNN-Com-Scabpit

# In[65]:


path_list = []

for root, dirs, files in os.walk(path_Disease, topdown=False):
    for path in dirs:
        path_list.append(path)
        
wb = openpyxl.load_workbook(path_Disease+'/'+path+'/6.Alpha.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list)
    for folder in path_list:
        data_temp = pd.read_excel(path_Disease+'/'+folder+'/6.Alpha.xlsx', sheet_name=sheet_name)
        temp_df[folder] = data_temp.iloc[4][1:].values

    results_dic[sheet_name] = temp_df


# In[66]:


main_df2 = pd.DataFrame(columns=['Data', 'Response'])
for key in results_dic.keys():
    for col in results_dic[key]:
        temp_df = pd.DataFrame()
        temp_df['Data'] = results_dic[key][col].values
        temp_df['Response'] = col
        temp_df['Level'] = key
        main_df2= pd.concat([main_df2, temp_df]) 


# In[67]:


main_df2['Method']='Alpha'


# In[68]:


path_list = []

for root, dirs, files in os.walk(path_Disease, topdown=False):
    for path in dirs:
        path_list.append(path)
        
wb = openpyxl.load_workbook(path_Disease+'/'+path+'/5.OTU-Score3.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list)
    for folder in path_list:
        data_temp = pd.read_excel(path_Disease+'/'+folder+'/5.OTU-Score3.xlsx', sheet_name=sheet_name)
        temp_df[folder] = data_temp.iloc[4][1:].values

    results_dic[sheet_name] = temp_df


# In[69]:


main_df3 = pd.DataFrame(columns=['Data', 'Response'])
for key in results_dic.keys():
    for col in results_dic[key]:
        temp_df = pd.DataFrame()
        temp_df['Data'] = results_dic[key][col].values
        temp_df['Response'] = col
        temp_df['Level'] = key
        main_df3= pd.concat([main_df3, temp_df]) 


# In[70]:


main_df3['Method']='OTU-S3'


# In[71]:


Main2 = pd.concat([main_df2,main_df3],axis=0)


# In[72]:


path_list = []

for root, dirs, files in os.walk(path_Disease, topdown=False):
    for path in dirs:
        path_list.append(path)
        
wb = openpyxl.load_workbook(path_Disease+'/'+path+'/92.Alpha+Soil+Disease.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list)
    for folder in path_list:
        data_temp = pd.read_excel(path_Disease+'/'+folder+'/92.Alpha+Soil+Disease.xlsx', sheet_name=sheet_name)
        temp_df[folder] = data_temp.iloc[4][1:].values

    results_dic[sheet_name] = temp_df


# In[73]:


main_df4 = pd.DataFrame(columns=['Data', 'Response'])
for key in results_dic.keys():
    for col in results_dic[key]:
        temp_df = pd.DataFrame()
        temp_df['Data'] = results_dic[key][col].values
        temp_df['Response'] = col
        temp_df['Level'] = key
        main_df4= pd.concat([main_df4, temp_df]) 


# In[74]:


main_df4['Method']='Alpha+Soil+DS'


# In[75]:


Main3 = pd.concat([Main2,main_df4],axis=0)


# In[76]:


path_list = []

for root, dirs, files in os.walk(path_Disease, topdown=False):
    for path in dirs:
        path_list.append(path)
        
wb = openpyxl.load_workbook(path_Disease+'/'+path+'/95.OTU-Score3+Soil+Disease.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list)
    for folder in path_list:
        data_temp = pd.read_excel(path_Disease+'/'+folder+'/95.OTU-Score3+Soil+Disease.xlsx', sheet_name=sheet_name)
        temp_df[folder] = data_temp.iloc[4][1:].values

    results_dic[sheet_name] = temp_df


# In[77]:


main_df5 = pd.DataFrame(columns=['Data', 'Response'])
for key in results_dic.keys():
    for col in results_dic[key]:
        temp_df = pd.DataFrame()
        temp_df['Data'] = results_dic[key][col].values
        temp_df['Response'] = col
        temp_df['Level'] = key
        main_df5= pd.concat([main_df5, temp_df]) 


# In[78]:


main_df5['Method']='OTU-S3+Soil+DS'


# In[79]:


Main4 = pd.concat([Main3,main_df5],axis=0)


# In[80]:


Main4=Main4[Main4['Response']=='Scabpit']


# BNN

# In[81]:


path_list = []

for root, dirs, files in os.walk(path_Disease_BNN, topdown=False):
    for path in dirs:
        path_list.append(path)
        
wb = openpyxl.load_workbook(path_Disease_BNN+'/'+path+'/6.Alpha.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list)
    for folder in path_list:
        data_temp = pd.read_excel(path_Disease_BNN+'/'+folder+'/6.Alpha.xlsx', sheet_name=sheet_name)
        temp_df[folder] = data_temp.iloc[4][1:].values

    results_dic[sheet_name] = temp_df


# In[82]:


main_df2 = pd.DataFrame(columns=['Data', 'Response'])
for key in results_dic.keys():
    for col in results_dic[key]:
        temp_df = pd.DataFrame()
        temp_df['Data'] = results_dic[key][col].values
        temp_df['Response'] = col
        temp_df['Level'] = key
        main_df2= pd.concat([main_df2, temp_df]) 


# In[83]:


main_df2['Method']='Alpha'


# In[84]:


path_list = []

for root, dirs, files in os.walk(path_Disease_BNN, topdown=False):
    for path in dirs:
        path_list.append(path)
        
wb = openpyxl.load_workbook(path_Disease_BNN+'/'+path+'/5.OTU-Score3.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list)
    for folder in path_list:
        data_temp = pd.read_excel(path_Disease_BNN+'/'+folder+'/5.OTU-Score3.xlsx', sheet_name=sheet_name)
        temp_df[folder] = data_temp.iloc[4][1:].values

    results_dic[sheet_name] = temp_df


# In[85]:


main_df3 = pd.DataFrame(columns=['Data', 'Response'])
for key in results_dic.keys():
    for col in results_dic[key]:
        temp_df = pd.DataFrame()
        temp_df['Data'] = results_dic[key][col].values
        temp_df['Response'] = col
        temp_df['Level'] = key
        main_df3= pd.concat([main_df3, temp_df]) 


# In[86]:


main_df3['Method']='OTU-S3'


# In[87]:


Main2 = pd.concat([main_df2,main_df3],axis=0)


# In[88]:


path_list = []

for root, dirs, files in os.walk(path_Disease_BNN, topdown=False):
    for path in dirs:
        path_list.append(path)
        
wb = openpyxl.load_workbook(path_Disease_BNN+'/'+path+'/92.Alpha+Soil+Disease.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list)
    for folder in path_list:
        data_temp = pd.read_excel(path_Disease_BNN+'/'+folder+'/92.Alpha+Soil+Disease.xlsx', sheet_name=sheet_name)
        temp_df[folder] = data_temp.iloc[4][1:].values

    results_dic[sheet_name] = temp_df


# In[89]:


main_df4 = pd.DataFrame(columns=['Data', 'Response'])
for key in results_dic.keys():
    for col in results_dic[key]:
        temp_df = pd.DataFrame()
        temp_df['Data'] = results_dic[key][col].values
        temp_df['Response'] = col
        temp_df['Level'] = key
        main_df4= pd.concat([main_df4, temp_df]) 


# In[90]:


main_df4['Method']='Alpha+Soil+DS'


# In[91]:


Main3 = pd.concat([Main2,main_df4],axis=0)


# In[92]:


path_list = []

for root, dirs, files in os.walk(path_Disease_BNN, topdown=False):
    for path in dirs:
        path_list.append(path)
        
wb = openpyxl.load_workbook(path_Disease_BNN+'/'+path+'/95.OTU-Score3+Soil+Disease.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list)
    for folder in path_list:
        data_temp = pd.read_excel(path_Disease_BNN+'/'+folder+'/95.OTU-Score3+Soil+Disease.xlsx', sheet_name=sheet_name)
        temp_df[folder] = data_temp.iloc[4][1:].values

    results_dic[sheet_name] = temp_df


# In[93]:


main_df5 = pd.DataFrame(columns=['Data', 'Response'])
for key in results_dic.keys():
    for col in results_dic[key]:
        temp_df = pd.DataFrame()
        temp_df['Data'] = results_dic[key][col].values
        temp_df['Response'] = col
        temp_df['Level'] = key
        main_df5= pd.concat([main_df5, temp_df]) 


# In[94]:


main_df5['Method']='OTU-S3+Soil+DS'


# In[95]:


Main42 = pd.concat([Main3,main_df5],axis=0)


# In[96]:


Main44=Main42[Main42['Response']=='Scabpit']


# In[97]:


fig, axs = plt.subplots(1,2, figsize=(20,5), facecolor='w', edgecolor='k')
order=['Alpha','Alpha+Soil+DS','OTU-S3','OTU-S3+Soil+DS']
plt.subplot(1,2,1)
AA = sns.boxplot(x="Method", hue="Level", y="Data", data=Main4,palette='Set2',hue_order =hue_order ,order=order)
plt.xticks(fontsize=18)
plt.yticks(fontsize=15)
AA.grid(axis= 'y')
plt.legend([],[], frameon=False)
plt.ylabel('Weighted F1-score',fontsize=16)
plt.xlabel('')
plt.title('Random Forest',fontsize=20)
plt.ylim(0,1)

plt.subplot(1,2,2)
AA2 = sns.boxplot(x="Method", hue="Level", y="Data", data=Main44,palette='Set2',hue_order =hue_order,order=order )
#plt.xticks(fontweight='bold')
plt.ylabel('')
plt.xlabel('')
plt.xticks(fontsize=18)
plt.title('Bayesian NN',fontsize=20)
ax = axs[0]

plt.ylim(0,1)

legend = plt.legend(loc='lower right',fontsize=13,mode = "expand", ncol = 5)

AA2.axes.yaxis.set_ticklabels([])
AA2.grid(axis= 'y')
fig.subplots_adjust(hspace =0, wspace=0)

plt.savefig(path_figure+'15-RF-BNN-Com-Scabpit.png',bbox_inches = 'tight',dpi=600,facecolor='white')





