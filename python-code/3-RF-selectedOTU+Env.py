#!/usr/bin/env python
# coding: utf-8

# In[20]:


import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.model_selection import RepeatedKFold
from sklearn.pipeline import Pipeline
import os
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.metrics import confusion_matrix, classification_report
from sklearn.model_selection import GridSearchCV
from pandas import ExcelWriter
from sklearn.ensemble import RandomForestClassifier
import openpyxl
from sklearn.preprocessing import StandardScaler
from sklearn.preprocessing import MinMaxScaler


# In[21]:


from sklearn.preprocessing import OneHotEncoder
from sklearn.compose import make_column_transformer
from sklearn.pipeline import make_pipeline


# In[22]:


import warnings
warnings.filterwarnings('ignore')


# In[23]:


from datetime import datetime
start_time = datetime.now()


# In[1]:


list_level = ['Class', 'Family', 'Genus', 'Order', 'Phylum']
path_response='response/response_original/'
data_path = 'response/response_netcomi/'
path_Count='OTU/Count_data/'
path_soil = 'Env/soil_chemistry/'
path_disease = 'Env/disease_suppression/'
path_field='Env/field_information/'


# In[24]:


def process_data(data_train,data_val,cv):  
    x_column_list = data_train.drop(columns=['y_b']).columns
    #Classification RF  
    pipeRF = Pipeline([('classifier', [RandomForestClassifier()])])  
    param_grid = [  
    {'classifier' : [RandomForestClassifier()], 
    'classifier__n_estimators': [100, 200],
    'classifier__min_samples_split': [8, 10],
    'classifier__min_samples_leaf': [3, 4, 5],
     'classifier__max_depth': [80, 90],
    'classifier__criterion':('gini','entropy'),  
    'classifier__class_weight':('balanced','auto')}]  
    clf = GridSearchCV(pipeRF, param_grid = param_grid, cv = cv, n_jobs=-1, scoring='f1_weighted')  
    # Fit on data  
    clf.fit(data_train[x_column_list],data_train['y_b'])  
    best_clf=clf.best_estimator_  
    y_valid=best_clf.predict(data_val[x_column_list])  
    report_All = classification_report(data_val['y_b'],y_valid,output_dict=True)  
    dAll=pd.DataFrame(report_All).transpose()  
    return dAll


# In[25]:


cv=RepeatedKFold(n_splits=10,n_repeats=3, random_state=100)


# In[27]:


path_list2 = []
#reading files in folder response
for root, dirs, files in os.walk(path_response, topdown=False):
    for path in dirs:
        path_list2.append(path)
#reading sheet name       
wb = openpyxl.load_workbook(path_response+path+'/feature_selection.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list2, index=range(0,700))
    for folder in path_list2:
        data_temp = pd.read_excel(path_response+folder+'/feature_selection.xlsx', sheet_name=sheet_name)
        temp_df[folder].iloc[
            range(0, len(data_temp['Unnamed: 0'].values))] = data_temp['Unnamed: 0'].values
    if sheet_name =='Phylum':
        temp_df = temp_df.iloc[range(0,30)]
    elif sheet_name =='Class':
        temp_df = temp_df.iloc[range(0,36)]
    elif sheet_name =='Order':
        temp_df = temp_df.iloc[range(0,75)]
    elif sheet_name =='Family':
        temp_df = temp_df.iloc[range(0,85)]
    elif sheet_name =='Genus':
        temp_df = temp_df.iloc[range(0,162)]
        
    results_dic[sheet_name] = temp_df
    print( results_dic[sheet_name].shape)


# In[28]:


path_list = []

for root, dirs, files in os.walk(data_path, topdown=False):
    for path in dirs:
        path_list.append(path)


# In[29]:


all_data = dict.fromkeys(dirs)
for folder in path_list:
    
    all_data[folder] = dict.fromkeys(list_level)
    temp_path = data_path+folder
    temp_file_list = os.listdir(temp_path)
    for file in temp_file_list:
        if 'Count' in file:
            level = file.split('_')[-1][:-4]
            data_temp = pd.read_csv(data_path+folder+'/'+file,index_col=0)
            data_temp.sort_values(by='abs.diff..x', ascending=False, inplace=True)
            all_data[folder][level] = data_temp.iloc[range(0,max(30, round(len(data_temp)/3)))]
print('Done.')


# In[30]:


SELECTED_FEATURE = dict.fromkeys(dirs)
for col in range(0,len(list_level)):
    level = list_level[col]
    print(level)
    SELECTED_FEATURE[level] = dict.fromkeys(dirs)
    response_list = path_list
    feature_list = []
    for response in response_list:
        feature_list.append(all_data[response][level].index.union(results_dic[level][response][results_dic[level][response].notnull()].values))
    print(all_data[response][level].index.shape)
    print(results_dic[level][response].shape)
    feature_list = [item for subitem in feature_list for item in subitem]
    feature_list = np.unique(feature_list)
    print(len(feature_list))
    matrix_df = pd.DataFrame(columns = response_list, index = feature_list)
    for response in response_list:
        for feature in feature_list:
                if (feature in all_data[response][level].index) & (feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 3
                elif(feature in all_data[response][level].index) &(feature not in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 2##NetComi
                elif (feature not in all_data[response][level].index) &(feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 1##ML
                else:
                    matrix_df[response].loc[feature] = 0##NotMLnotNetcomi
   
        SELECTED_FEATURE[level][response] = matrix_df[response][matrix_df[response]==2].index
        print(SELECTED_FEATURE[level][response].shape[0])
    matrix_df['Sum'] = 0
    for index in matrix_df.index:
        matrix_df['Sum'].loc[index] = sum(matrix_df.loc[index].values)
    
    matrix_df.sort_values(by='Sum', ascending=False, inplace=True)
    


# In[31]:


for file_response in os.listdir(path_response):  
    if (file_response != '.DS_Store') & (file_response != 'Icon\r'):  
        print(file_response)  
        path_r= path_response+file_response
        for re in os.listdir(path_r):  
            if re[0:8] == 'response':  
                response = pd.read_csv(path_r+'/'+re)
                response.rename(columns={'Column1':'Link_ID',response.columns[1]:'y_b'}, inplace=True)  
                path_x = path_Count  
                writer= pd.ExcelWriter(path_r+'/'+'classification_RF_FS_Netcomi'+'.xlsx', engine='xlsxwriter')   
                for file_folder in os.listdir(path_x):  
                    if (file_folder[-4:] != '.csv') & (file_folder != '.DS_Store')& (file_folder != 'Icon\r'):          
                        path = path_x+file_folder
                        file_list = []  
                        tRF=pd.DataFrame()
                        k=0  
                        for file in os.listdir(path):  
                            if (file[0] != 't') & (file[-4:] == '.csv') & (file != '.DS_Store')& (file_folder != 'Icon\r'):  
                                print(file)  
                                file_list.append(file)
                                data_temp = pd.read_csv(path+'/'+file)
                                data_temp.rename(columns={'Unnamed: 0':'Link_ID'}, inplace=True) 
                                A = [item for item in SELECTED_FEATURE[file_folder][file_response]]
                                A.insert(0, "Link_ID")
                                data_temp = data_temp[A]
                                data_temp.rename(columns={'Unnamed: 0':'Link_ID'}, inplace=True)  
                                RD1 = data_temp.drop('Link_ID',axis=1)
                                RD11 = RD1.div(RD1.sum(axis=1), axis=0)
                                data_temp = pd.concat([data_temp['Link_ID'],RD11],axis=1)                            
                                data=pd.merge(response,data_temp,on='Link_ID')
                                data.drop(columns = 'Link_ID',inplace=True)  
                                data.dropna(inplace=True)
                                data_train,data_val = train_test_split(data,train_size=0.8, random_state=42)  
                                output = process_data(data_train,data_val,cv)  
                                tRF[k]=pd.DataFrame(output['f1-score'].values) 
                                k=k+1           
                        tRF.to_excel(writer, sheet_name=file_folder, index=True) 
                writer.save()


# # ML result

# In[32]:


SELECTED_FEATURE = dict.fromkeys(dirs)
for col in range(0,len(list_level)):
    level = list_level[col]
    print(level)
    SELECTED_FEATURE[level] = dict.fromkeys(dirs)
    response_list = path_list
    feature_list = []
    for response in response_list:
        feature_list.append(all_data[response][level].index.union(results_dic[level][response][results_dic[level][response].notnull()].values))
    feature_list = [item for subitem in feature_list for item in subitem]
    feature_list = np.unique(feature_list)
    print(len(feature_list))
    matrix_df = pd.DataFrame(columns = response_list, index = feature_list)
    for response in response_list:
        for feature in feature_list:
                if (feature in all_data[response][level].index) & (feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 3
                elif(feature in all_data[response][level].index) &(feature not in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 2##NetComi
                elif (feature not in all_data[response][level].index) &(feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 1##ML
                else:
                    matrix_df[response].loc[feature] = 0##NotMLnotNetcomi
        SELECTED_FEATURE[level][response] = matrix_df[response][matrix_df[response]==1].index
        print(SELECTED_FEATURE[level][response].shape[0])
    matrix_df['Sum'] = 0
    for index in matrix_df.index:
        matrix_df['Sum'].loc[index] = sum(matrix_df.loc[index].values)
    
    matrix_df.sort_values(by='Sum', ascending=False, inplace=True)


# In[33]:


for file_response in os.listdir(path_response):  
    if (file_response != '.DS_Store') & (file_response != 'Icon\r'):  
        print(file_response)  
        path_r= path_response+file_response
        for re in os.listdir(path_r):  
            if re[0:8] == 'response':  
                response = pd.read_csv(path_r+'/'+re)
                response.rename(columns={'Column1':'Link_ID',response.columns[1]:'y_b'}, inplace=True)  
                path_x = path_Count  
                writer= pd.ExcelWriter(path_r+'/'+'classification_RF_FS_ML'+'.xlsx', engine='xlsxwriter')   
                for file_folder in os.listdir(path_x):  
                    if (file_folder[-4:] != '.csv') & (file_folder != '.DS_Store')& (file_folder != 'Icon\r'):          
                        path = path_x+file_folder
                        file_list = []  
                        tRF=pd.DataFrame() 
                        tcluster=pd.DataFrame()  
                        k=0  
                        for file in os.listdir(path):  
                            if (file[0] != 't') & (file[-4:] == '.csv') & (file != '.DS_Store')& (file_folder != 'Icon\r'):  
                                print(file)                                  
                                 
                                file_list.append(file)  
                                data_temp = pd.read_csv(path+'/'+file)
                                data_temp.rename(columns={'Unnamed: 0':'Link_ID'}, inplace=True) 
                                A = [item for item in SELECTED_FEATURE[file_folder][file_response]]
                                A.insert(0, "Link_ID")
                                data_temp = data_temp[A]
                                data_temp.rename(columns={'Unnamed: 0':'Link_ID'}, inplace=True)  
                                RD1 = data_temp.drop('Link_ID',axis=1)
                                RD11 = RD1.div(RD1.sum(axis=1), axis=0)
                                data_temp = pd.concat([data_temp['Link_ID'],RD11],axis=1)
                                
                                data=pd.merge(response,data_temp,on='Link_ID')
                                data.drop(columns = 'Link_ID',inplace=True)  
                                data.dropna(inplace=True)
                                data_train,data_val = train_test_split(data,train_size=0.8, random_state=42)  
                                output = process_data(data_train,data_val,cv)  
                                tRF[k]=pd.DataFrame(output['f1-score'].values) 
                                k=k+1           
                        tRF.to_excel(writer, sheet_name=file_folder, index=True) 
                writer.save()


# # important OTU (ML&Netcomi)Score==3

# In[34]:


SELECTED_FEATURE = dict.fromkeys(dirs)
for col in range(0,len(list_level)):
    level = list_level[col]
    print(level)
    SELECTED_FEATURE[level] = dict.fromkeys(dirs)
    response_list = path_list
    feature_list = []
    for response in response_list:
        feature_list.append(all_data[response][level].index.union(results_dic[level][response][results_dic[level][response].notnull()].values))
    feature_list = [item for subitem in feature_list for item in subitem]
    feature_list = np.unique(feature_list)
    print(len(feature_list))
    matrix_df = pd.DataFrame(columns = response_list, index = feature_list)
    for response in response_list:
        for feature in feature_list:
                if (feature in all_data[response][level].index) & (feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 3
                elif(feature in all_data[response][level].index) &(feature not in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 2##NetComi
                elif (feature not in all_data[response][level].index) &(feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 1##ML
                else:
                    matrix_df[response].loc[feature] = 0##NotMLnotNetcomi
        SELECTED_FEATURE[level][response] = matrix_df[response][matrix_df[response]==3].index
    
    matrix_df['Sum'] = 0
    for index in matrix_df.index:
        matrix_df['Sum'].loc[index] = sum(matrix_df.loc[index].values)
    
    matrix_df.sort_values(by='Sum', ascending=False, inplace=True)


# In[35]:


for file_response in os.listdir(path_response):  
    if (file_response != '.DS_Store') & (file_response != 'Icon\r'):  
        print(file_response)  
        path_r= path_response+file_response
        for re in os.listdir(path_r):  
            if re[0:8] == 'response':  
                response = pd.read_csv(path_r+'/'+re)
                response.rename(columns={'Column1':'Link_ID',response.columns[1]:'y_b'}, inplace=True)  
                path_x = path_Count
                writer= pd.ExcelWriter(path_r+'/'+'classification_RF_FS_score3'+'.xlsx', engine='xlsxwriter')   
                for file_folder in os.listdir(path_x):  
                    if (file_folder[-4:] != '.csv') & (file_folder != '.DS_Store')& (file_folder != 'Icon\r'):          
                        path = path_x+file_folder
                        file_list = []  
                        tRF=pd.DataFrame()  
                        tcluster=pd.DataFrame()  
                        k=0  
                        for file in os.listdir(path):  
                            if (file[0] != 't') & (file[-4:] == '.csv') & (file != '.DS_Store')& (file_folder != 'Icon\r'):  
                                print(file)  
                                file_list.append(file)  
                                data_temp = pd.read_csv(path+'/'+file)
                                data_temp.rename(columns={'Unnamed: 0':'Link_ID'}, inplace=True) 
                                A = [item for item in SELECTED_FEATURE[file_folder][file_response]]
                                A.insert(0, "Link_ID")
                                data_temp = data_temp[A]
                                data_temp.rename(columns={'Unnamed: 0':'Link_ID'}, inplace=True)  
                                RD1 = data_temp.drop('Link_ID',axis=1)
                                RD11 = RD1.div(RD1.sum(axis=1), axis=0)
                                data_temp = pd.concat([data_temp['Link_ID'],RD11],axis=1)        
                                data=pd.merge(response,data_temp,on='Link_ID')
                                data.drop(columns = 'Link_ID',inplace=True)  
                                data.dropna(inplace=True)
                                data_train,data_val = train_test_split(data,train_size=0.8, random_state=42)  
                                output = process_data(data_train,data_val,cv)  
                                tRF[k]=pd.DataFrame(output['f1-score'].values) 
                                k=k+1
                                
                        tRF.to_excel(writer, sheet_name=file_folder, index=True)  
                writer.save()                      


# # Not important (Score0) size=size _score3

# In[36]:


SELECTED_FEATURE = dict.fromkeys(dirs)
SELECTED_FEATURE_0 = dict.fromkeys(dirs)
for col in range(0,len(list_level)):
    level = list_level[col]
    print(level)
    SELECTED_FEATURE[level] = dict.fromkeys(dirs)
    SELECTED_FEATURE_0[level] = dict.fromkeys(dirs)
    response_list = path_list
    feature_list = []
    for response in response_list:
        feature_list.append(all_data[response][level].index.union(results_dic[level][response][results_dic[level][response].notnull()].values))
    feature_list = [item for subitem in feature_list for item in subitem]
    feature_list = np.unique(feature_list)
    print(len(feature_list))
    matrix_df = pd.DataFrame(columns = response_list, index = feature_list)
    for response in response_list:
        for feature in feature_list:
                if (feature in all_data[response][level].index) & (feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 3
                elif(feature in all_data[response][level].index) &(feature not in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 2##NetComi
                elif (feature not in all_data[response][level].index) &(feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 1##ML
                else:
                    matrix_df[response].loc[feature] = 0##NotMLnotNetcomi   
        SELECTED_FEATURE[level][response] = matrix_df[response][matrix_df[response]==3].index
        print(SELECTED_FEATURE[level][response].shape[0])
        SELECTED_FEATURE_0[level][response] = matrix_df[response][matrix_df[response]==0].index[0:SELECTED_FEATURE[level][response].shape[0]] 
    matrix_df['Sum'] = 0
    for index in matrix_df.index:
        matrix_df['Sum'].loc[index] = sum(matrix_df.loc[index].values)
    matrix_df.sort_values(by='Sum', ascending=False, inplace=True)


# In[37]:


for file_response in os.listdir(path_response):  
    if (file_response != '.DS_Store') & (file_response != 'Icon\r'):  
        print(file_response)  
        path_r= path_response+file_response
        for re in os.listdir(path_r):  
            if re[0:8] == 'response':  
                response = pd.read_csv(path_r+'/'+re)
                response.rename(columns={'Column1':'Link_ID',response.columns[1]:'y_b'}, inplace=True)   
                path_x = path_Count
                writer= pd.ExcelWriter(path_r+'/'+'classification_RF_FS_notImportant'+'.xlsx', engine='xlsxwriter')   
                for file_folder in os.listdir(path_x):  
                    if (file_folder[-4:] != '.csv') & (file_folder != '.DS_Store')& (file_folder != 'Icon\r'):          
                        path = path_x+file_folder
                        file_list = []  
                        tRF=pd.DataFrame()  
                        tcluster=pd.DataFrame()  
                        k=0  
                        for file in os.listdir(path):  
                            if (file[0] != 't') & (file[-4:] == '.csv') & (file != '.DS_Store')& (file_folder != 'Icon\r'):  
                                file_list.append(file)  
                                data_temp = pd.read_csv(path+'/'+file)
                                data_temp.rename(columns={'Unnamed: 0':'Link_ID'}, inplace=True) 
                                A = [item for item in SELECTED_FEATURE_0[file_folder][file_response]]
                                A.insert(0, "Link_ID")
                                data_temp = data_temp[A]
                                data_temp.rename(columns={'Unnamed: 0':'Link_ID'}, inplace=True)  
                                RD1 = data_temp.drop('Link_ID',axis=1)
                                RD11 = RD1.div(RD1.sum(axis=1), axis=0)
                                data_temp = pd.concat([data_temp['Link_ID'],RD11],axis=1)
                                data=pd.merge(response,data_temp,on='Link_ID')
                                data.drop(columns = 'Link_ID',inplace=True)  
                                data.dropna(inplace=True)
                                data_train,data_val = train_test_split(data,train_size=0.8, random_state=42)  
                                output = process_data(data_train,data_val,cv)  
                                tRF[k]=pd.DataFrame(output['f1-score'].values) 
                                k=k+1                
                        tRF.to_excel(writer, sheet_name=file_folder, index=True)  
                writer.save()                      


# # disease_suppression+important features

# In[38]:


path_list2 = []
#reading files in folder response
for root, dirs, files in os.walk(path_response, topdown=False):
    for path in dirs:
        path_list2.append(path)
#reading sheet name       
wb = openpyxl.load_workbook(path_response+path+'/feature_selection.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list2, index=range(0,700))
    for folder in path_list2:
        data_temp = pd.read_excel(path_response+folder+'/feature_selection.xlsx', sheet_name=sheet_name)
        temp_df[folder].iloc[
            range(0, len(data_temp['Unnamed: 0'].values))] = data_temp['Unnamed: 0'].values
    if sheet_name =='Phylum':
        temp_df = temp_df.iloc[range(0,30)]
    elif sheet_name =='Class':
        temp_df = temp_df.iloc[range(0,36)]
    elif sheet_name =='Order':
        temp_df = temp_df.iloc[range(0,75)]
    elif sheet_name =='Family':
        temp_df = temp_df.iloc[range(0,85)]
    elif sheet_name =='Genus':
        temp_df = temp_df.iloc[range(0,162)]
        
    results_dic[sheet_name] = temp_df
    print( results_dic[sheet_name].shape)


# In[39]:


SELECTED_FEATURE = dict.fromkeys(dirs) 
for col in range(0,len(list_level)):
    level = list_level[col]
    print(level)
    SELECTED_FEATURE[level] = dict.fromkeys(dirs)
    response_list = path_list
    feature_list = []
    for response in response_list:
        feature_list.append(all_data[response][level].index.union(results_dic[level][response][results_dic[level][response].notnull()].values))
    print(all_data[response][level].index.shape)
    print(results_dic[level][response].shape)
    feature_list = [item for subitem in feature_list for item in subitem]
    feature_list = np.unique(feature_list)
    print(len(feature_list))
    matrix_df = pd.DataFrame(columns = response_list, index = feature_list)
    for response in response_list:
        for feature in feature_list:
                if (feature in all_data[response][level].index) & (feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 3
                elif(feature in all_data[response][level].index) &(feature not in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 2##NetComi
                elif (feature not in all_data[response][level].index) &(feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 1##ML
                else:
                    matrix_df[response].loc[feature] = 0##NotMLnotNetcomi   
        SELECTED_FEATURE[level][response] = matrix_df[response][matrix_df[response]==3].index
        print(SELECTED_FEATURE[level][response].shape[0])
    matrix_df['Sum'] = 0
    for index in matrix_df.index:
        matrix_df['Sum'].loc[index] = sum(matrix_df.loc[index].values)
    
    matrix_df.sort_values(by='Sum', ascending=False, inplace=True)
    matrix_df.to_excel(writer, sheet_name=level, index=True)
writer.save()


# In[40]:


for file_response in os.listdir(path_response):  
    if (file_response != '.DS_Store') & (file_response != 'Icon\r'):  
        print(file_response)  
        path_r= path_response+file_response
        for re in os.listdir(path_r):  
            if re[0:8] == 'response':  
                response = pd.read_csv(path_r+'/'+re)
                response.rename(columns={'Column1':'Link_ID',response.columns[1]:'y_b'}, inplace=True)  
                path_x = path_Count
                writer= pd.ExcelWriter(path_r+'/'+'classification_RF_FS_disease_suppression'+'.xlsx', engine='xlsxwriter')   
                for file_folder in os.listdir(path_x):  
                    if (file_folder[-4:] != '.csv') & (file_folder != '.DS_Store')& (file_folder != 'Icon\r'):          
                        path = path_x+file_folder
                        file_list = []  
                        tRF=pd.DataFrame()   
                        k=0  
                        for file in os.listdir(path):  
                            if (file[0] != 't') & (file[-4:] == '.csv') & (file != '.DS_Store')& (file_folder != 'Icon\r'):  
                                print(file)  
                                file_list.append(file)  
                                data_temp = pd.read_csv(path+'/'+file)
                                data_temp.rename(columns={'Unnamed: 0':'Link_ID'}, inplace=True) 
                                A = [item for item in SELECTED_FEATURE[file_folder][file_response]]
                                A.insert(0, "Link_ID")
                                data_temp = data_temp[A]
                                data_temp.rename(columns={'Unnamed: 0':'Link_ID'}, inplace=True) 
                                RD1 = data_temp.drop('Link_ID',axis=1)
                                RD11 = RD1.div(RD1.sum(axis=1), axis=0)
                                data_temp_original = pd.concat([data_temp['Link_ID'],RD11],axis=1)
                                
                                
                                
                                for k in range(1,7):
                                    fe = pd.read_csv(path_disease+str(k)+'.csv')
                                    fe = fe.drop('Unnamed: 0',axis=1)
                                    data_temp = pd.merge(data_temp_original,fe,on='Link_ID')

                                    data=pd.merge(response,data_temp,on='Link_ID')
                                    data.dropna(inplace=True)
                                
                                data.drop(columns = 'Link_ID',inplace=True)  
                                data_train,data_val = train_test_split(data,train_size=0.8, random_state=42)  
                                output = process_data(data_train,data_val,cv)  
                                tRF[k]=pd.DataFrame(output['f1-score'].values) 
                                k=k+1  
                        tRF.to_excel(writer, sheet_name=file_folder, index=True) 
                writer.save()


# # soil_chemistry+important features

# In[41]:


path_list2 = []
#reading files in folder response
for root, dirs, files in os.walk(path_response, topdown=False):
    for path in dirs:
        path_list2.append(path)
#reading sheet name       
wb = openpyxl.load_workbook(path_response+path+'/feature_selection.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list2, index=range(0,700))
    for folder in path_list2:
        data_temp = pd.read_excel(path_response+folder+'/feature_selection.xlsx', sheet_name=sheet_name)
        temp_df[folder].iloc[
            range(0, len(data_temp['Unnamed: 0'].values))] = data_temp['Unnamed: 0'].values
    if sheet_name =='Phylum':
        temp_df = temp_df.iloc[range(0,30)]
    elif sheet_name =='Class':
        temp_df = temp_df.iloc[range(0,36)]
    elif sheet_name =='Order':
        temp_df = temp_df.iloc[range(0,75)]
    elif sheet_name =='Family':
        temp_df = temp_df.iloc[range(0,85)]
    elif sheet_name =='Genus':
        temp_df = temp_df.iloc[range(0,162)]
        
    results_dic[sheet_name] = temp_df
    print( results_dic[sheet_name].shape)


# In[42]:


SELECTED_FEATURE = dict.fromkeys(dirs)
for col in range(0,len(list_level)):
    level = list_level[col]
    print(level)
    SELECTED_FEATURE[level] = dict.fromkeys(dirs)
    response_list = path_list
    feature_list = []
    for response in response_list:
        feature_list.append(all_data[response][level].index.union(results_dic[level][response][results_dic[level][response].notnull()].values))
    print(all_data[response][level].index.shape)
    print(results_dic[level][response].shape)
    feature_list = [item for subitem in feature_list for item in subitem]
    feature_list = np.unique(feature_list)
    print(len(feature_list))
    matrix_df = pd.DataFrame(columns = response_list, index = feature_list)
    for response in response_list:
        for feature in feature_list:
                if (feature in all_data[response][level].index) & (feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 3
                elif(feature in all_data[response][level].index) &(feature not in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 2##NetComi
                elif (feature not in all_data[response][level].index) &(feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 1##ML
                else:
                    matrix_df[response].loc[feature] = 0##NotMLnotNetcomi 
        SELECTED_FEATURE[level][response] = matrix_df[response][matrix_df[response]==3].index
        print(SELECTED_FEATURE[level][response].shape[0])
    matrix_df['Sum'] = 0
    for index in matrix_df.index:
        matrix_df['Sum'].loc[index] = sum(matrix_df.loc[index].values)
    
    matrix_df.sort_values(by='Sum', ascending=False, inplace=True)


# In[43]:


for file_response in os.listdir(path_response):  
    if (file_response != '.DS_Store') & (file_response != 'Icon\r'):  
        print(file_response)  
        path_r= path_response+file_response
        for re in os.listdir(path_r):  
            if re[0:8] == 'response':  
                response = pd.read_csv(path_r+'/'+re)
                response.rename(columns={'Column1':'Link_ID',response.columns[1]:'y_b'}, inplace=True)  
                path_x = path_Count
                writer= pd.ExcelWriter(path_r+'/'+'classification_RF_FS_soil_chemistry'+'.xlsx', engine='xlsxwriter')
                for file_folder in os.listdir(path_x):  
                    if (file_folder[-4:] != '.csv') & (file_folder != '.DS_Store')& (file_folder != 'Icon\r'):          
                        path = path_x+file_folder
                        file_list = []  
                        tRF=pd.DataFrame()  
                        tcluster=pd.DataFrame()  
                        #k=0  
                        for file in os.listdir(path):  
                            if (file[0] != 't') & (file[-4:] == '.csv') & (file != '.DS_Store')& (file != 'Icon\r'):  
                                print(file)  
                                file_list.append(file)  
                                data_temp = pd.read_csv(path+'/'+file)
                                data_temp.rename(columns={'Unnamed: 0':'Link_ID'}, inplace=True) 
                                A = [item for item in SELECTED_FEATURE[file_folder][file_response]]
                                A.insert(0, "Link_ID")
                                data_temp = data_temp[A]
                                data_temp.rename(columns={'Unnamed: 0':'Link_ID'}, inplace=True) 
                                RD1 = data_temp.drop('Link_ID',axis=1)
                                RD11 = RD1.div(RD1.sum(axis=1), axis=0)
                                data_temp_original = pd.concat([data_temp['Link_ID'],RD11],axis=1)
                                
                                
                                
                                for k in range(1,7):
                                    fe = pd.read_csv(path_soil+str(k)+'.csv')
                                    fe = fe.drop('Unnamed: 0',axis=1)
                                    data_temp = pd.merge(data_temp_original,fe,on='Link_ID')

                                    data=pd.merge(response,data_temp,on='Link_ID')
                                    data.dropna(inplace=True)

                                    data.drop(columns = 'Link_ID',inplace=True)  
                                    data_train,data_val = train_test_split(data,train_size=0.8, random_state=42)  
                                    output = process_data(data_train,data_val,cv)  
                                    tRF[k]=pd.DataFrame(output['f1-score'].values) 
                                    
                        tRF.to_excel(writer, sheet_name=file_folder, index=True)  
                writer.save() 


# # Field_information+important features

# In[44]:


path_list2 = []
#reading files in folder response
for root, dirs, files in os.walk(path_response, topdown=False):
    for path in dirs:
        path_list2.append(path)
#reading sheet name       
wb = openpyxl.load_workbook(path_response+path+'/feature_selection.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list2, index=range(0,700))
    for folder in path_list2:
        data_temp = pd.read_excel(path_response+folder+'/feature_selection.xlsx', sheet_name=sheet_name)
        temp_df[folder].iloc[
            range(0, len(data_temp['Unnamed: 0'].values))] = data_temp['Unnamed: 0'].values
    if sheet_name =='Phylum':
        temp_df = temp_df.iloc[range(0,30)]
    elif sheet_name =='Class':
        temp_df = temp_df.iloc[range(0,36)]
    elif sheet_name =='Order':
        temp_df = temp_df.iloc[range(0,75)]
    elif sheet_name =='Family':
        temp_df = temp_df.iloc[range(0,85)]
    elif sheet_name =='Genus':
        temp_df = temp_df.iloc[range(0,162)]
        
    results_dic[sheet_name] = temp_df
    print( results_dic[sheet_name].shape)


# In[45]:


SELECTED_FEATURE = dict.fromkeys(dirs)
for col in range(0,len(list_level)):
    level = list_level[col]
    print(level)
    SELECTED_FEATURE[level] = dict.fromkeys(dirs)
    response_list = path_list
    feature_list = []
    for response in response_list:
        feature_list.append(all_data[response][level].index.union(results_dic[level][response][results_dic[level][response].notnull()].values))
    print(all_data[response][level].index.shape)
    print(results_dic[level][response].shape)
    feature_list = [item for subitem in feature_list for item in subitem]
    feature_list = np.unique(feature_list)
    print(len(feature_list))
    matrix_df = pd.DataFrame(columns = response_list, index = feature_list)
    for response in response_list:
        for feature in feature_list:
                if (feature in all_data[response][level].index) & (feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 3
                elif(feature in all_data[response][level].index) &(feature not in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 2##NetComi
                elif (feature not in all_data[response][level].index) &(feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 1##ML
                else:
                    matrix_df[response].loc[feature] = 0##NotMLnotNetcomi   
        SELECTED_FEATURE[level][response] = matrix_df[response][matrix_df[response]==3].index
        print(SELECTED_FEATURE[level][response].shape[0])
    matrix_df['Sum'] = 0
    for index in matrix_df.index:
        matrix_df['Sum'].loc[index] = sum(matrix_df.loc[index].values)
    
    matrix_df.sort_values(by='Sum', ascending=False, inplace=True)


# In[46]:


rf_classifier = RandomForestClassifier(
min_samples_leaf=50,
n_estimators=150,
bootstrap=True,
oob_score=True,
n_jobs=-1,
random_state=42,
max_features='auto')


# In[48]:


for file_response in os.listdir(path_response):  
    if (file_response != '.DS_Store') & (file_response != 'Icon\r'):  
        print(file_response)  
        path_r= path_response+file_response
        for re in os.listdir(path_r):  
            if re[0:8] == 'response':  
                response = pd.read_csv(path_r+'/'+re)
                response.rename(columns={'Column1':'Link_ID',response.columns[1]:'y_b'}, inplace=True)  
                path_x = path_Count
                writer= pd.ExcelWriter(path_r+'/'+'classification_RF_FS_field_information'+'.xlsx', engine='xlsxwriter')
                for file_folder in os.listdir(path_x):  
                    if (file_folder[-4:] != '.csv') & (file_folder != '.DS_Store')& (file_folder != 'Icon\r'):          
                        path = path_x+file_folder
                        file_list = []  
                        tRF=pd.DataFrame() 
                        tcluster=pd.DataFrame()  
                        k=0  
                        for file in os.listdir(path):  
                            if (file[0] != 't') & (file[-4:] == '.csv') & (file != '.DS_Store')& (file_folder != 'Icon\r'):  
                                print(file)  
                                file_list.append(file)  
                                data_temp = pd.read_csv(path+'/'+file)
                                data_temp.rename(columns={'Unnamed: 0':'Link_ID'}, inplace=True) 
                                A = [item for item in SELECTED_FEATURE[file_folder][file_response]]
                                A.insert(0, "Link_ID")
                                data_temp = data_temp[A]
                                data_temp.rename(columns={'Unnamed: 0':'Link_ID'}, inplace=True) 
                                RD1 = data_temp.drop('Link_ID',axis=1)
                                RD11 = RD1.div(RD1.sum(axis=1), axis=0)
                                data_temp = pd.concat([data_temp['Link_ID'],RD11],axis=1)
                                
                                
                                fe=pd.read_csv(path_field+'field_information.csv')
                                fe.rename(columns={'Column1':'Link_ID'},inplace=True)
                                A = fe.columns[fe.isna().sum()< 30]

                                fe = fe[A]
                                
                                level_cols = fe.columns 
                                level_cols =level_cols.drop(['Link_ID','Variety'])
                                
                                features_to_encode = level_cols.drop(['pH_1_1'])
                                
                                col_trans = make_column_transformer(
                                (OneHotEncoder(),features_to_encode),
                                remainder = "passthrough")
                                df_scaled = fe[level_cols]
                                fe_data=pd.concat([fe['Link_ID'],df_scaled],axis=1)
                                data_temp = pd.merge(data_temp,fe_data,on='Link_ID')
                                
                                data=pd.merge(response,data_temp,on='Link_ID')
                                data.drop(columns = ['Link_ID','pH_1_1'],inplace=True)
                                data.dropna(inplace=True)
                                AL = []
                                for l in range(0,50):
                                    data_train,data_val = train_test_split(data,train_size=0.97, random_state=l)


                                    x_column_list = data_train.drop(columns=['y_b']).columns  
                                    #Classification RF 

                                    pipe = make_pipeline(col_trans, rf_classifier)
                                    # Fit on data  
                                    pipe.fit(data_train[x_column_list],data_train['y_b'])  
                                    y_valid = pipe.predict(data_val[x_column_list])

                                    report_All = classification_report(data_val['y_b'],y_valid,output_dict=True)  
                                    dAll=pd.DataFrame(report_All).transpose() 
                                    AL.append(dAll['f1-score'].values)
                                AAA=pd.DataFrame(AL)
                                tRF[k]=AAA.mean()  

                                k=k+1  
                        tRF.to_excel(writer, sheet_name=file_folder, index=True) 
                        
                writer.save()  


# # important_otu+soil+disease+field

# In[49]:


fe=pd.read_csv(path_field+'field_information.csv')
fe.rename(columns={"Column1":'Link_ID'},inplace=True)
A = fe.columns[fe.isna().sum()< 30]
fe = fe[A]
level_cols = fe.columns 
level_cols =level_cols.drop(['Link_ID','Variety'])
features_to_encode = level_cols.drop(['pH_1_1'])
col_trans = make_column_transformer((OneHotEncoder(),features_to_encode),remainder = "passthrough")

rf_classifier = RandomForestClassifier(
min_samples_leaf=50,
n_estimators=150,
bootstrap=True,
oob_score=True,
n_jobs=-1,
random_state=42,
max_features='auto')


# In[50]:



path_list2 = []
#reading files in folder response
for root, dirs, files in os.walk(path_response, topdown=False):
    for path in dirs:
        path_list2.append(path)
#reading sheet name       
wb = openpyxl.load_workbook(path_response+path+'/feature_selection.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list2, index=range(0,700))
    for folder in path_list2:
        data_temp = pd.read_excel(path_response+folder+'/feature_selection.xlsx', sheet_name=sheet_name)
        temp_df[folder].iloc[
            range(0, len(data_temp['Unnamed: 0'].values))] = data_temp['Unnamed: 0'].values
    if sheet_name =='Phylum':
        temp_df = temp_df.iloc[range(0,30)]
    elif sheet_name =='Class':
        temp_df = temp_df.iloc[range(0,36)]
    elif sheet_name =='Order':
        temp_df = temp_df.iloc[range(0,75)]
    elif sheet_name =='Family':
        temp_df = temp_df.iloc[range(0,85)]
    elif sheet_name =='Genus':
        temp_df = temp_df.iloc[range(0,162)]
        
    results_dic[sheet_name] = temp_df
    print( results_dic[sheet_name].shape)


# In[51]:


SELECTED_FEATURE = dict.fromkeys(dirs)
for col in range(0,len(list_level)):
    level = list_level[col]
    print(level)
    SELECTED_FEATURE[level] = dict.fromkeys(dirs)
    response_list = path_list
    feature_list = []
    for response in response_list:
        feature_list.append(all_data[response][level].index.union(results_dic[level][response][results_dic[level][response].notnull()].values))
    print(all_data[response][level].index.shape)
    print(results_dic[level][response].shape)
    feature_list = [item for subitem in feature_list for item in subitem]
    feature_list = np.unique(feature_list)
    print(len(feature_list))
    matrix_df = pd.DataFrame(columns = response_list, index = feature_list)
    for response in response_list:
        for feature in feature_list:
                if (feature in all_data[response][level].index) & (feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 3
                elif(feature in all_data[response][level].index) &(feature not in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 2##NetComi
                elif (feature not in all_data[response][level].index) &(feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 1##ML
                else:
                    matrix_df[response].loc[feature] = 0##NotMLnotNetcomi
        SELECTED_FEATURE[level][response] = matrix_df[response][matrix_df[response]==3].index
        print(SELECTED_FEATURE[level][response].shape[0])
    matrix_df['Sum'] = 0
    for index in matrix_df.index:
        matrix_df['Sum'].loc[index] = sum(matrix_df.loc[index].values)
    
    matrix_df.sort_values(by='Sum', ascending=False, inplace=True)


# In[52]:


for file_response in os.listdir(path_response):  
    if (file_response != '.DS_Store') & (file_response != 'Icon\r'):  
        print(file_response)  
        path_r= path_response+file_response
        for re in os.listdir(path_r):  
            if re[0:8] == 'response':  
                response = pd.read_csv(path_r+'/'+re)
                response.rename(columns={'Column1':'Link_ID',response.columns[1]:'y_b'}, inplace=True)     
                path_x = path_Count
                writer= pd.ExcelWriter(path_r+'/'+'classification_important_otu+soil+disease+field'+'.xlsx', engine='xlsxwriter')
    
                tRF=pd.DataFrame()    
                k=0  
                for level in os.listdir(path_x):
                    if (level[-4:] != '.csv') &  (level != '.DS_Store'):
                        for k in range(1,7):
                            f2 = pd.read_csv(path_soil+str(k)+'.csv')
                            f3 = pd.read_csv(path_disease+str(k)+'.csv')
                            f23=pd.merge(f2,f3,on='Link_ID')
                            f23.drop(columns=['Unnamed: 0_x','Unnamed: 0_y'],inplace=True)
                            data1 = pd.merge(response,f23,on='Link_ID')
                            fe_data=pd.concat([fe['Link_ID'],fe[level_cols]],axis=1)
                            
                            data = pd.merge(data1,fe_data,on='Link_ID')
                            data.dropna(inplace=True)
                            print(data.shape)
                            data['Variety2'] = data['Variety2'].replace("RedLittle","Red")
                            
                            
                            path = path_x+level
                            for file in os.listdir(path):  
                                if (file[0] != 't') & (file[-4:] == '.csv') & (file != '.DS_Store')& (file != 'Icon\r'):  
                                    print(file)    
                                    data_temp = pd.read_csv(path+'/'+file)
                                    
                                    data_temp.rename(columns={'Unnamed: 0':'Link_ID'}, inplace=True) 
                                    A = [item for item in SELECTED_FEATURE[level][file_response]]
                                    A.insert(0, "Link_ID")
                                    data_temp = data_temp[A]
                                    data_temp.rename(columns={'Unnamed: 0':'Link_ID'}, inplace=True) 
                                    RD1 = data_temp.drop('Link_ID',axis=1)
                                    RD11 = RD1.div(RD1.sum(axis=1), axis=0)
                                    data_temp = pd.concat([data_temp['Link_ID'],RD11],axis=1)
                                    data_2 = pd.merge(data_temp,data,on='Link_ID')                        
                                    data_2.drop(columns = 'Link_ID',inplace=True) 
                                    data_2.dropna(inplace=True)
                                    AL = []
                                    for l in range(0,100):
                                        data_train,data_val = train_test_split(data_2,train_size=0.95, random_state=l)


                                        x_column_list = data_train.drop(columns=['y_b']).columns  
                                        #Classification RF 

                                        pipe = make_pipeline(col_trans, rf_classifier)
                                        # Fit on data  
                                        pipe.fit(data_train[x_column_list],data_train['y_b'])  
                                        y_valid = pipe.predict(data_val[x_column_list])

                                        report_All = classification_report(data_val['y_b'],y_valid,output_dict=True)  
                                        dAll=pd.DataFrame(report_All).transpose() 
                                        AL.append(dAll['f1-score'].values)
                                    AAA=pd.DataFrame(AL)
                                    tRF[k]=AAA.mean() 
                                    
                                    k=k+1  
                        tRF.to_excel(writer, sheet_name=level, index=True) 
        writer.save() 


# # important_otu+soil+disease

# In[53]:


path_list2 = []
#reading files in folder response
for root, dirs, files in os.walk(path_response, topdown=False):
    for path in dirs:
        path_list2.append(path)
#reading sheet name       
wb = openpyxl.load_workbook(path_response+path+'/feature_selection.xlsx')
sheet_list = wb.sheetnames

results_dic = dict.fromkeys(sheet_list)

for sheet_name in results_dic.keys():
    temp_df = pd.DataFrame(columns=path_list2, index=range(0,700))
    for folder in path_list2:
        data_temp = pd.read_excel(path_response+folder+'/feature_selection.xlsx', sheet_name=sheet_name)
        temp_df[folder].iloc[
            range(0, len(data_temp['Unnamed: 0'].values))] = data_temp['Unnamed: 0'].values
    if sheet_name =='Phylum':
        temp_df = temp_df.iloc[range(0,30)]
    elif sheet_name =='Class':
        temp_df = temp_df.iloc[range(0,36)]
    elif sheet_name =='Order':
        temp_df = temp_df.iloc[range(0,75)]
    elif sheet_name =='Family':
        temp_df = temp_df.iloc[range(0,85)]
    elif sheet_name =='Genus':
        temp_df = temp_df.iloc[range(0,162)]
        
    results_dic[sheet_name] = temp_df
    print( results_dic[sheet_name].shape)


# In[54]:


SELECTED_FEATURE = dict.fromkeys(dirs)
for col in range(0,len(list_level)):
    level = list_level[col]
    print(level)
    SELECTED_FEATURE[level] = dict.fromkeys(dirs)
    response_list = path_list
    feature_list = []
    for response in response_list:
        feature_list.append(all_data[response][level].index.union(results_dic[level][response][results_dic[level][response].notnull()].values))
    print(all_data[response][level].index.shape)
    print(results_dic[level][response].shape)
    feature_list = [item for subitem in feature_list for item in subitem]
    feature_list = np.unique(feature_list)
    print(len(feature_list))
    matrix_df = pd.DataFrame(columns = response_list, index = feature_list)
    for response in response_list:
        for feature in feature_list:
                if (feature in all_data[response][level].index) & (feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 3
                elif(feature in all_data[response][level].index) &(feature not in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 2##NetComi
                elif (feature not in all_data[response][level].index) &(feature in results_dic[level][response].values):
                    matrix_df[response].loc[feature] = 1##ML
                else:
                    matrix_df[response].loc[feature] = 0##NotMLnotNetcomi   
        SELECTED_FEATURE[level][response] = matrix_df[response][matrix_df[response]==3].index
        print(SELECTED_FEATURE[level][response].shape[0])
    matrix_df['Sum'] = 0
    for index in matrix_df.index:
        matrix_df['Sum'].loc[index] = sum(matrix_df.loc[index].values)
    
    matrix_df.sort_values(by='Sum', ascending=False, inplace=True)


# In[ ]:


end_time = datetime.now()
print('Duration: {}'.format(end_time - start_time))

